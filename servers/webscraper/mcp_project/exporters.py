"""
Export system — Convert scraped results to multiple output formats.

Supports: JSON, CSV, Excel (.xlsx), SQLite, Markdown tables.
"""

import csv
import io
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional


def _flatten_record(record: dict, prefix: str = "") -> dict:
    """Flatten nested dicts for tabular export."""
    flat = {}
    for k, v in record.items():
        key = f"{prefix}{k}" if not prefix else f"{prefix}.{k}"
        if isinstance(v, dict):
            flat.update(_flatten_record(v, key))
        elif isinstance(v, list):
            flat[key] = "; ".join(str(i) for i in v)
        else:
            flat[key] = v
    return flat


def export_json(results: list[dict], output_path: str) -> str:
    """Export results as formatted JSON."""
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump({
            "exported_at": datetime.utcnow().isoformat() + "Z",
            "total_records": len(results),
            "data": results,
        }, f, indent=2, default=str, ensure_ascii=False)
    return output_path


def export_csv(results: list[dict], output_path: str) -> str:
    """Export results as CSV with flattened columns."""
    if not results:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("")
        return output_path

    # Flatten all records
    flat_records = [_flatten_record(r) for r in results]

    # Collect all unique keys (preserving order from first record)
    all_keys = []
    seen = set()
    for rec in flat_records:
        for k in rec:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(flat_records)

    return output_path


def export_excel(results: list[dict], output_path: str) -> str:
    """Export results as Excel workbook with data sheet and metadata sheet."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Data Sheet ──
    ws = wb.active
    ws.title = "Scraped Data"

    if not results:
        ws["A1"] = "No data"
        wb.save(output_path)
        return output_path

    flat_records = [_flatten_record(r) for r in results]

    # Collect columns
    all_keys = []
    seen = set()
    for rec in flat_records:
        for k in rec:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    # Header style
    header_fill = PatternFill(start_color="2D4A3E", end_color="2D4A3E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, rec in enumerate(flat_records, 2):
        for col_idx, key in enumerate(all_keys, 1):
            value = rec.get(key, "")
            if isinstance(value, (dict, list)):
                value = str(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width columns (capped at 50)
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Metadata Sheet ──
    ws_meta = wb.create_sheet("Export Info")
    meta_data = [
        ("Export Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Total Records", len(results)),
        ("Columns", len(all_keys)),
        ("Source", "MCP Web Scraper"),
    ]
    for row, (k, v) in enumerate(meta_data, 1):
        ws_meta.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws_meta.cell(row=row, column=2, value=str(v))

    wb.save(output_path)
    return output_path


def export_sqlite(results: list[dict], output_path: str, table_name: str = "scraped_data") -> str:
    """Export results to a SQLite database."""
    import sqlite3

    if not results:
        conn = sqlite3.connect(output_path)
        conn.execute(f"CREATE TABLE IF NOT EXISTS {table_name} (id INTEGER PRIMARY KEY)")
        conn.close()
        return output_path

    flat_records = [_flatten_record(r) for r in results]

    all_keys = []
    seen = set()
    for rec in flat_records:
        for k in rec:
            if k not in seen:
                # Sanitize column names for SQL
                safe_k = k.replace(".", "_").replace("-", "_").replace(" ", "_")
                all_keys.append((k, safe_k))
                seen.add(k)

    conn = sqlite3.connect(output_path)
    cursor = conn.cursor()

    # Create table
    cols = ", ".join(f'"{safe}" TEXT' for _, safe in all_keys)
    cursor.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols})')

    # Insert data
    placeholders = ", ".join(["?"] * len(all_keys))
    col_names = ", ".join(f'"{safe}"' for _, safe in all_keys)
    for rec in flat_records:
        values = [str(rec.get(orig, "")) for orig, _ in all_keys]
        cursor.execute(f'INSERT INTO "{table_name}" ({col_names}) VALUES ({placeholders})', values)

    conn.commit()
    conn.close()
    return output_path


def export_markdown(results: list[dict], output_path: str) -> str:
    """Export results as Markdown table."""
    if not results:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("# Scraped Data\n\n_No data_\n")
        return output_path

    flat_records = [_flatten_record(r) for r in results]

    all_keys = []
    seen = set()
    for rec in flat_records:
        for k in rec:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    lines = [f"# Scraped Data\n", f"_Exported {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}_\n"]

    # Header
    lines.append("| " + " | ".join(all_keys) + " |")
    lines.append("| " + " | ".join(["---"] * len(all_keys)) + " |")

    # Rows
    for rec in flat_records:
        row = []
        for k in all_keys:
            val = str(rec.get(k, "")).replace("|", "\\|").replace("\n", " ")
            if len(val) > 100:
                val = val[:97] + "..."
            row.append(val)
        lines.append("| " + " | ".join(row) + " |")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    return output_path


EXPORT_FORMATS = {
    "json": {"extension": ".json", "function": export_json, "description": "Structured JSON with metadata"},
    "csv": {"extension": ".csv", "function": export_csv, "description": "Comma-separated values (flat)"},
    "excel": {"extension": ".xlsx", "function": export_excel, "description": "Excel workbook with formatting"},
    "sqlite": {"extension": ".db", "function": export_sqlite, "description": "SQLite database"},
    "markdown": {"extension": ".md", "function": export_markdown, "description": "Markdown table"},
}


def export_results(
    results: list[dict],
    format: str,
    output_dir: str,
    filename: Optional[str] = None,
) -> str:
    """
    Export results to the specified format.
    
    Returns the path to the exported file.
    """
    if format not in EXPORT_FORMATS:
        raise ValueError(f"Unknown format: {format}. Available: {list(EXPORT_FORMATS.keys())}")

    fmt = EXPORT_FORMATS[format]
    if not filename:
        filename = f"scrape_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(output_dir / f"{filename}{fmt['extension']}")

    return fmt["function"](results, output_path)
