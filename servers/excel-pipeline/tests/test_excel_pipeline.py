"""
Tests for the Excel Pipeline MCP server.

Covers: tool contracts, schema conformance, error paths, readiness states.
"""

from __future__ import annotations

import json
import os
import sys
import tempfile
from pathlib import Path

import openpyxl
import pytest

# Ensure the mcp_project package is importable
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "mcp_project"))

import services
import tools
from models import TriageResult, ValidationResult


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def sample_workbook(tmp_path: Path) -> str:
    """Create a simple well-structured workbook for testing."""
    fp = tmp_path / "test_workbook.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plant Summary"

    # Header row
    ws.append(["Asset", "Installed Capacity (MW)", "Net Output (GWh)", "Availability (%)"])
    # Data rows
    ws.append(["Alpha Plant", 250.0, 1200.5, 92.3])
    ws.append(["Beta Plant", 180.0, 850.0, 88.1])
    ws.append(["Gamma Plant", 320.0, 1500.0, 95.0])

    wb.save(str(fp))
    return str(fp)


@pytest.fixture()
def multi_table_workbook(tmp_path: Path) -> str:
    """Create a workbook with multiple tables separated by blank rows."""
    fp = tmp_path / "multi_table.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Table 1
    ws.append(["Region", "Revenue", "Costs"])
    ws.append(["North", 1000, 600])
    ws.append(["South", 800, 500])

    # Blank separator
    ws.append([None, None, None])
    ws.append([None, None, None])

    # Table 2
    ws.append(["Product", "Units Sold", "Price"])
    ws.append(["Widget", 500, 9.99])
    ws.append(["Gadget", 300, 19.99])

    wb.save(str(fp))
    return str(fp)


@pytest.fixture()
def empty_workbook(tmp_path: Path) -> str:
    """Create a workbook with no meaningful data."""
    fp = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(fp))
    return str(fp)


@pytest.fixture(autouse=True)
def _clear_state():
    """Reset in-memory stores between tests."""
    services._workbooks.clear()
    services._sheets.clear()
    services._detected_tables.clear()
    services._canonical_models.clear()
    services._lineage_store.clear()
    services._bundles.clear()


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _parse(json_str: str) -> dict:
    return json.loads(json_str)


# ===========================================================================
# 1. triage_workbook
# ===========================================================================

class TestTriageWorkbook:
    def test_supported_workbook(self, sample_workbook):
        res = _parse(tools.triage_workbook(sample_workbook))
        assert res["status"] == "success"
        assert res["triage_result"] in [t.value for t in TriageResult]
        assert "summary" in res
        assert res["summary"]["sheet_count"] >= 1
        assert "lineage_ref" in res

    def test_file_not_found(self):
        res = _parse(tools.triage_workbook("/nonexistent/path.xlsx"))
        assert res["status"] == "error"
        assert res["error"]["code"] == "FILE_NOT_FOUND"

    def test_unsupported_extension(self, tmp_path):
        fp = tmp_path / "test.txt"
        fp.write_text("not an excel file")
        res = _parse(tools.triage_workbook(str(fp)))
        assert res["status"] == "error"
        assert res["error"]["code"] == "UNSUPPORTED_FILE_TYPE"


# ===========================================================================
# 2. inspect_workbook_structure
# ===========================================================================

class TestInspectWorkbookStructure:
    def test_returns_sheets(self, sample_workbook):
        res = _parse(tools.inspect_workbook_structure(sample_workbook))
        assert res["status"] == "success"
        assert len(res["sheets"]) >= 1
        assert res["sheets"][0]["sheet_name"] == "Plant Summary"

    def test_lineage_ref_present(self, sample_workbook):
        res = _parse(tools.inspect_workbook_structure(sample_workbook))
        assert "lineage_ref" in res
        assert "workbook_id" in res["lineage_ref"]


# ===========================================================================
# 3. detect_candidate_tables
# ===========================================================================

class TestDetectCandidateTables:
    def test_detects_single_table(self, sample_workbook):
        res = _parse(tools.detect_candidate_tables(sample_workbook))
        assert res["status"] == "success"
        assert res["summary"]["table_count"] >= 1
        tbl = res["detected_tables"][0]
        assert "detection_confidence" in tbl
        assert tbl["detection_confidence"] > 0.5

    def test_detects_multiple_tables(self, multi_table_workbook):
        res = _parse(tools.detect_candidate_tables(multi_table_workbook))
        assert res["status"] == "success"
        assert res["summary"]["table_count"] >= 2

    def test_specific_sheet(self, sample_workbook):
        res = _parse(tools.detect_candidate_tables(sample_workbook, sheet_names=["Plant Summary"]))
        assert res["status"] == "success"
        assert res["summary"]["table_count"] >= 1


# ===========================================================================
# 4. build_canonical_model
# ===========================================================================

class TestBuildCanonicalModel:
    def test_builds_model(self, sample_workbook):
        res = _parse(tools.build_canonical_model(sample_workbook))
        assert res["status"] == "success"
        entities = res["canonical_model"]["entities"]
        assert entities["assets"] >= 1
        assert entities["kpis"] >= 1
        assert "model_id" in res["lineage_ref"]

    def test_model_has_measurements(self, sample_workbook):
        res = _parse(tools.build_canonical_model(sample_workbook))
        assert res["canonical_model"]["entities"]["measurements"] >= 1


# ===========================================================================
# 5. validate_canonical_model
# ===========================================================================

class TestValidateCanonicalModel:
    def test_validates_successfully(self, sample_workbook):
        build = _parse(tools.build_canonical_model(sample_workbook))
        model_id = build["lineage_ref"]["model_id"]
        res = _parse(tools.validate_canonical_model(model_id))
        assert res["status"] == "success"
        assert res["validation_result"] in [v.value for v in ValidationResult]
        assert "confidence_summary" in res

    def test_unknown_model_id(self):
        res = _parse(tools.validate_canonical_model("nonexistent"))
        assert res["status"] == "error"
        assert res["error"]["code"] == "VALIDATION_BLOCKER"


# ===========================================================================
# 6. get_lineage
# ===========================================================================

class TestGetLineage:
    def test_lineage_found(self, sample_workbook):
        build = _parse(tools.build_canonical_model(sample_workbook))
        model_id = build["lineage_ref"]["model_id"]
        model = services._canonical_models[model_id]
        if model.measurements:
            m = model.measurements[0]
            res = _parse(tools.get_lineage("Measurement", m.measurement_id))
            assert res["status"] == "success"
            assert "lineage" in res

    def test_lineage_not_found(self):
        res = _parse(tools.get_lineage("Asset", "nonexistent"))
        assert res["status"] == "error"


# ===========================================================================
# 7. export_ready_json
# ===========================================================================

class TestExportReadyJson:
    def test_export_after_validation(self, sample_workbook, tmp_path):
        build = _parse(tools.build_canonical_model(sample_workbook))
        model_id = build["lineage_ref"]["model_id"]
        _parse(tools.validate_canonical_model(model_id))
        # Override output dir for test isolation
        services._canonical_models[model_id].ready_for_export = True
        result = services.export_ready_json(model_id, output_dir=str(tmp_path / "exports"))
        assert "export" in result
        assert os.path.exists(result["export"]["uri"])

    def test_export_blocked_without_validation(self, sample_workbook):
        build = _parse(tools.build_canonical_model(sample_workbook))
        model_id = build["lineage_ref"]["model_id"]
        # Don't validate → ready_for_export is False
        res = _parse(tools.export_ready_json(model_id))
        assert res["status"] == "error"
        assert res["error"]["code"] == "EXPORT_BLOCKED"


# ===========================================================================
# 8. prepare_figure_specs
# ===========================================================================

class TestPrepareFigureSpecs:
    def test_generates_specs(self, sample_workbook):
        build = _parse(tools.build_canonical_model(sample_workbook))
        model_id = build["lineage_ref"]["model_id"]
        res = _parse(tools.prepare_figure_specs(model_id))
        assert res["status"] == "success"
        assert "figure_specs" in res


# ===========================================================================
# 9. create_report_bundle
# ===========================================================================

class TestCreateReportBundle:
    def test_creates_bundle(self, sample_workbook, tmp_path):
        build = _parse(tools.build_canonical_model(sample_workbook))
        model_id = build["lineage_ref"]["model_id"]
        _parse(tools.validate_canonical_model(model_id))
        services._canonical_models[model_id].ready_for_export = True
        result = services.create_report_bundle(model_id, output_dir=str(tmp_path / "bundles"))
        assert "bundle" in result
        assert len(result["bundle"]["contents"]) >= 1


# ===========================================================================
# 10. grounded_query_validated_data
# ===========================================================================

class TestGroundedQuery:
    def test_query_highest(self, sample_workbook):
        build = _parse(tools.build_canonical_model(sample_workbook))
        model_id = build["lineage_ref"]["model_id"]
        res = _parse(tools.grounded_query_validated_data(model_id, "Which asset has the highest value?"))
        assert res["status"] == "success"
        assert "answer" in res
        assert res["answer"]["confidence"] > 0

    def test_query_count(self, sample_workbook):
        build = _parse(tools.build_canonical_model(sample_workbook))
        model_id = build["lineage_ref"]["model_id"]
        res = _parse(tools.grounded_query_validated_data(model_id, "How many assets are there?"))
        assert res["status"] == "success"


# ===========================================================================
# 11. process_workbook_end_to_end
# ===========================================================================

class TestEndToEnd:
    def test_full_pipeline(self, sample_workbook):
        res = _parse(tools.process_workbook_end_to_end(sample_workbook))
        assert res["status"] == "success"
        assert "model_id" in res

    def test_unsupported_file(self, tmp_path):
        fp = tmp_path / "bad.csv"
        fp.write_text("a,b,c")
        res = _parse(tools.process_workbook_end_to_end(str(fp)))
        assert res["status"] == "error"


# ===========================================================================
# 12. Error contract conformance
# ===========================================================================

class TestErrorContract:
    """Ensure all error responses follow the standard envelope."""

    def test_error_has_code_and_message(self):
        res = _parse(tools.triage_workbook("/nonexistent.xlsx"))
        assert "error" in res
        assert "code" in res["error"]
        assert "message" in res["error"]
        assert "severity" in res["error"]
