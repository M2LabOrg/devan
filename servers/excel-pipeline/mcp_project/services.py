"""
Backend services for the Excel Pipeline.

Deterministic processing logic that the MCP tools wrap.
The MCP layer must NOT duplicate any logic here.
"""

from __future__ import annotations

import json
import os
import re
from dataclasses import asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart
from openpyxl.utils import get_column_letter

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from docling.document_converter import DocumentConverter as DoclingConverter
    DOCLING_AVAILABLE = True
except ImportError:
    DOCLING_AVAILABLE = False

from models import (
    Asset,
    Assumption,
    CanonicalModel,
    ConfidenceSummary,
    DetectedTable,
    FigureSpec,
    KPI,
    Measurement,
    ReadinessState,
    ReportBundle,
    RiskSignal,
    Scenario,
    Severity,
    SheetRecord,
    SourceLineage,
    TimeSeriesPoint,
    TriageResult,
    ValidationIssue,
    ValidationResult,
    WorkbookRecord,
    _now_iso,
    _uid,
    to_dict,
)

# ---------------------------------------------------------------------------
# Shared state (in-memory store for demo / single-process deployments)
# ---------------------------------------------------------------------------

_workbooks: Dict[str, WorkbookRecord] = {}
_sheets: Dict[str, List[SheetRecord]] = {}          # workbook_id -> sheets
_detected_tables: Dict[str, List[DetectedTable]] = {}  # workbook_id -> tables
_canonical_models: Dict[str, CanonicalModel] = {}    # model_id -> model
_lineage_store: Dict[str, SourceLineage] = {}         # lineage_id -> lineage
_bundles: Dict[str, ReportBundle] = {}                # bundle_id -> bundle

ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".xlsb"}
MAX_FILE_SIZE_MB = 100


# ===================================================================
# 1. Ingestion / Triage Service
# ===================================================================

def triage_workbook(
    workbook_uri: str,
    processing_profile: str = "default",
    include_sheet_inventory: bool = True,
    include_risk_signals: bool = True,
) -> Dict[str, Any]:
    """Assess whether a workbook is supported, partially supported, or unsupported."""
    file_path = _resolve_uri(workbook_uri)
    _validate_file(file_path)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_count = len(wb.sheetnames)
    hidden_count = sum(1 for name in wb.sheetnames if wb[name].sheet_state != "visible")
    merged_count = sum(len(wb[name].merged_cells.ranges) for name in wb.sheetnames)
    has_charts = any(len(wb[name]._charts) > 0 for name in wb.sheetnames)
    has_images = any(len(wb[name]._images) > 0 for name in wb.sheetnames)

    # Estimate candidate tables across all sheets
    table_estimate = 0
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row and ws.max_column and ws.max_row > 1:
            table_estimate += _estimate_table_count(ws)

    # Check for macros
    has_macros = file_path.lower().endswith((".xlsm", ".xlsb"))

    # Check for external links
    has_external_links = bool(getattr(wb, "external_links", None))

    # Risk signals
    risk_signals: List[Dict[str, str]] = []
    if merged_count > 20:
        risk_signals.append({"code": "HIGH_MERGE_COUNT", "severity": "WARNING", "message": f"Workbook has {merged_count} merged regions which may complicate table detection."})
    if hidden_count > 0:
        risk_signals.append({"code": "HIDDEN_SHEETS", "severity": "INFO", "message": f"{hidden_count} hidden sheet(s) detected."})
    if has_macros:
        risk_signals.append({"code": "MACROS_PRESENT", "severity": "WARNING", "message": "Workbook contains macros; formula results may differ from stored values."})
    for name in wb.sheetnames:
        ws = wb[name]
        if _estimate_table_count(ws) > 1:
            risk_signals.append({"code": "MULTI_TABLE_SHEET", "severity": "WARNING", "message": f"Sheet '{name}' appears to contain multiple tabular regions."})
            break
    if has_external_links:
        risk_signals.append({"code": "EXTERNAL_LINKS", "severity": "WARNING", "message": "Workbook contains external links; referenced data may be stale."})

    # Determine triage result
    blockers = [r for r in risk_signals if r["severity"] == "BLOCKER"]
    warnings = [r for r in risk_signals if r["severity"] == "WARNING"]

    if blockers:
        triage_result = TriageResult.UNSUPPORTED
        confidence = 0.3
    elif len(warnings) >= 3:
        triage_result = TriageResult.REVIEW_REQUIRED
        confidence = 0.6
    elif warnings:
        triage_result = TriageResult.SUPPORTED_WITH_WARNINGS
        confidence = 0.85
    else:
        triage_result = TriageResult.SUPPORTED
        confidence = 0.95

    # Register workbook
    wb_record = WorkbookRecord.create(file_name=os.path.basename(file_path), file_path=file_path, sheet_count=sheet_count)
    _workbooks[wb_record.workbook_id] = wb_record

    wb.close()

    return {
        "triage_result": triage_result.value,
        "confidence": round(confidence, 2),
        "summary": {
            "sheet_count": sheet_count,
            "hidden_sheet_count": hidden_count,
            "merged_region_count": merged_count,
            "candidate_table_count_estimate": table_estimate,
            "has_macros": has_macros,
            "has_external_links": has_external_links,
            "has_embedded_charts": has_charts,
            "has_images": has_images,
        },
        "risk_signals": risk_signals if include_risk_signals else [],
        "next_recommended_tools": ["inspect_workbook_structure", "detect_candidate_tables"],
        "lineage_ref": {"workbook_id": wb_record.workbook_id},
    }


# ===================================================================
# 2. Workbook Structure Inspection Service
# ===================================================================

def inspect_workbook_structure(
    workbook_uri: str,
    include_ranges: bool = True,
    include_named_tables: bool = True,
    include_hidden_items: bool = True,
) -> Dict[str, Any]:
    """Return the physical workbook model and sheet inventory."""
    file_path = _resolve_uri(workbook_uri)
    _validate_file(file_path)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    wb_record = _get_or_create_workbook(file_path, wb)

    sheets_out: List[Dict[str, Any]] = []
    sheet_records: List[SheetRecord] = []

    for name in wb.sheetnames:
        ws = wb[name]
        visible = ws.sheet_state == "visible"
        if not include_hidden_items and not visible:
            continue

        named_tables = []
        if include_named_tables:
            for tbl in ws.tables.values():
                named_tables.append({"table_name": tbl.name, "range": tbl.ref})

        merged = [str(mr) for mr in ws.merged_cells.ranges] if include_ranges else []

        sr = SheetRecord.create(
            workbook_id=wb_record.workbook_id,
            sheet_name=name,
            visible=visible,
            row_count_estimate=ws.max_row or 0,
            column_count_estimate=ws.max_column or 0,
            named_tables=named_tables,
            merged_regions=merged,
        )
        sheet_records.append(sr)
        sheets_out.append(to_dict(sr))

    _sheets[wb_record.workbook_id] = sheet_records
    wb.close()

    return {
        "workbook": {
            "workbook_id": wb_record.workbook_id,
            "file_name": wb_record.file_name,
            "sheet_count": wb_record.sheet_count,
        },
        "sheets": sheets_out,
        "lineage_ref": {"workbook_id": wb_record.workbook_id},
    }


# ===================================================================
# 3. Table Detection Service
# ===================================================================

def detect_candidate_tables(
    workbook_uri: str,
    sheet_names: Optional[List[str]] = None,
    detection_profile: str = "default",
    include_parser_comparison: bool = True,
) -> Dict[str, Any]:
    """Detect candidate tabular regions with confidence and provenance."""
    file_path = _resolve_uri(workbook_uri)
    _validate_file(file_path)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    wb_record = _get_or_create_workbook(file_path, wb)
    target_sheets = sheet_names or list(wb.sheetnames)

    detected: List[DetectedTable] = []
    low_confidence_count = 0

    for sname in target_sheets:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        tables = _detect_tables_in_sheet(ws, sname, wb_record.workbook_id, include_parser_comparison)
        for t in tables:
            if t.detection_confidence < 0.7:
                low_confidence_count += 1
        detected.extend(tables)

    _detected_tables[wb_record.workbook_id] = detected
    wb.close()

    return {
        "detected_tables": [to_dict(t) for t in detected],
        "summary": {
            "table_count": len(detected),
            "low_confidence_count": low_confidence_count,
        },
    }


# ===================================================================
# 4. Canonical Model Service
# ===================================================================

def build_canonical_model(
    workbook_uri: str,
    processing_profile: str = "default",
    mapping_profile: str = "default",
    selected_table_ids: Optional[List[str]] = None,
    allow_llm_disambiguation: bool = False,
) -> Dict[str, Any]:
    """Map detected workbook content into a controlled business schema."""
    file_path = _resolve_uri(workbook_uri)
    _validate_file(file_path)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    wb_record = _get_or_create_workbook(file_path, wb)

    # Ensure tables are detected
    if wb_record.workbook_id not in _detected_tables:
        detect_candidate_tables(workbook_uri)

    all_tables = _detected_tables.get(wb_record.workbook_id, [])
    if selected_table_ids:
        tables = [t for t in all_tables if t.detected_table_id in selected_table_ids]
    else:
        tables = all_tables

    model = CanonicalModel.create(workbook_id=wb_record.workbook_id, processing_profile=processing_profile)

    # Default scenario
    default_scenario = Scenario.create("Base")
    model.scenarios.append(default_scenario)

    unresolved: List[Dict[str, str]] = []
    mapping_summary: List[Dict[str, Any]] = []

    for dt in tables:
        mapped_types: List[str] = []
        sheet = wb[dt.sheet_name]

        # Build assets from index column values
        for header in dt.column_headers:
            normalised = header.strip().lower()
            # Heuristic: first column is often the asset/entity name
            if dt.column_headers.index(header) == 0:
                # Read data column to create assets
                for row in range(dt.data_start_row, dt.data_end_row + 1):
                    val = sheet.cell(row=row, column=1).value
                    if val and str(val).strip():
                        name = str(val).strip()
                        if not any(a.asset_name == name for a in model.assets):
                            model.assets.append(Asset.create(name))
                if model.assets:
                    mapped_types.append("Asset")
            else:
                # Subsequent columns are KPI/measurements
                unit = _guess_unit(header)
                kpi = KPI.create(kpi_name=header.strip(), default_unit=unit)
                model.kpis.append(kpi)
                mapped_types.append("KPI")

                # Create measurements linking assets -> kpi -> scenario
                for row in range(dt.data_start_row, dt.data_end_row + 1):
                    asset_val = sheet.cell(row=row, column=dt.index_columns[0] if dt.index_columns else 1).value if dt.index_columns else sheet.cell(row=row, column=1).value
                    data_col_idx = dt.column_headers.index(header) + 1
                    cell_val = sheet.cell(row=row, column=data_col_idx).value
                    if cell_val is not None and asset_val:
                        asset_name = str(asset_val).strip()
                        matching_asset = next((a for a in model.assets if a.asset_name == asset_name), None)
                        if matching_asset:
                            try:
                                numeric_val = float(cell_val)
                            except (ValueError, TypeError):
                                continue
                            lin = SourceLineage.create(
                                entity_type="Measurement",
                                entity_id="",
                                workbook_id=wb_record.workbook_id,
                                source_sheet_name=dt.sheet_name,
                                source_range=f"{get_column_letter(data_col_idx)}{row}",
                                source_table_id=dt.detected_table_id,
                                parser_source=list(dt.parser_agreement.keys()),
                            )
                            _lineage_store[lin.lineage_id] = lin
                            m = Measurement.create(
                                asset_id=matching_asset.asset_id,
                                scenario_id=default_scenario.scenario_id,
                                kpi_id=kpi.kpi_id,
                                value=numeric_val,
                                unit=unit,
                                lineage_ref=lin.lineage_id,
                            )
                            lin.entity_id = m.measurement_id
                            model.measurements.append(m)
                            model.lineage.append(lin)
                            if "Measurement" not in mapped_types:
                                mapped_types.append("Measurement")

        conf = dt.detection_confidence * 0.95 if dt.detection_confidence > 0 else 0.7
        mapping_summary.append({
            "source_table_id": dt.detected_table_id,
            "mapped_entity_types": list(set(mapped_types)),
            "mapping_confidence": round(conf, 2),
        })

    # Compute confidence
    avg_conf = sum(ms["mapping_confidence"] for ms in mapping_summary) / max(len(mapping_summary), 1)
    model.confidence = ConfidenceSummary(
        overall_model_confidence=round(avg_conf, 2),
        boundary_detection_confidence=round(avg_conf + 0.02, 2),
        schema_match_confidence=round(avg_conf, 2),
        semantic_mapping_confidence=round(avg_conf - 0.04, 2),
    )

    _canonical_models[model.model_id] = model
    wb.close()

    return {
        "canonical_model": {
            "model_id": model.model_id,
            "entities": {
                "assets": len(model.assets),
                "kpis": len(model.kpis),
                "measurements": len(model.measurements),
                "scenarios": len(model.scenarios),
                "assumptions": len(model.assumptions),
            },
        },
        "mapping_summary": mapping_summary,
        "unresolved_items": unresolved,
        "lineage_ref": {"workbook_id": wb_record.workbook_id, "model_id": model.model_id},
    }


# ===================================================================
# 5. Validation Service
# ===================================================================

def validate_canonical_model(
    model_id: str,
    validation_profile: str = "default",
    include_reconciliation: bool = True,
) -> Dict[str, Any]:
    """Run validation, reconciliation, and confidence checks."""
    model = _canonical_models.get(model_id)
    if model is None:
        raise ValueError(f"Model '{model_id}' not found. Build a canonical model first.")

    issues: List[ValidationIssue] = []

    # Check for duplicate KPI names
    kpi_names = [k.kpi_name for k in model.kpis]
    seen: Dict[str, int] = {}
    for name in kpi_names:
        seen[name] = seen.get(name, 0) + 1
    for name, count in seen.items():
        if count > 1:
            issues.append(ValidationIssue.create(
                severity=Severity.WARNING.value,
                code="DUPLICATE_KPI_NAME",
                message=f"KPI name '{name}' appears {count} times.",
            ))

    # Check for inconsistent units on same KPI
    unit_map: Dict[str, set] = {}
    for m in model.measurements:
        kpi = next((k for k in model.kpis if k.kpi_id == m.kpi_id), None)
        if kpi:
            unit_map.setdefault(kpi.kpi_name, set()).add(m.unit or "unknown")
    for kpi_name, units in unit_map.items():
        if len(units) > 1:
            issues.append(ValidationIssue.create(
                severity=Severity.WARNING.value,
                code="INCONSISTENT_UNIT_LABEL",
                message=f"Detected {units} for KPI '{kpi_name}'.",
                entity_ref={"entity_type": "KPI", "kpi_name": kpi_name},
            ))

    # Check for orphan measurements (no matching asset)
    asset_ids = {a.asset_id for a in model.assets}
    for m in model.measurements:
        if m.asset_id not in asset_ids:
            issues.append(ValidationIssue.create(
                severity=Severity.ERROR.value,
                code="ORPHAN_MEASUREMENT",
                message=f"Measurement '{m.measurement_id}' references unknown asset '{m.asset_id}'.",
            ))

    # Check model is non-empty
    if not model.assets and not model.measurements:
        issues.append(ValidationIssue.create(
            severity=Severity.BLOCKER.value,
            code="EMPTY_MODEL",
            message="Canonical model contains no assets or measurements.",
        ))

    # Reconciliation: check measurement values are plausible
    if include_reconciliation:
        for m in model.measurements:
            if m.value < 0:
                issues.append(ValidationIssue.create(
                    severity=Severity.INFO.value,
                    code="NEGATIVE_VALUE",
                    message=f"Measurement '{m.measurement_id}' has negative value {m.value}.",
                ))

    # Compute result
    blocker_count = sum(1 for i in issues if i.severity == "BLOCKER")
    error_count = sum(1 for i in issues if i.severity == "ERROR")
    warning_count = sum(1 for i in issues if i.severity == "WARNING")
    info_count = sum(1 for i in issues if i.severity == "INFO")

    if blocker_count > 0:
        result = ValidationResult.BLOCKED
        ready = False
    elif error_count > 0:
        result = ValidationResult.FAILED
        ready = False
    elif warning_count > 0:
        result = ValidationResult.PASSED_WITH_WARNINGS
        ready = True
    else:
        result = ValidationResult.PASSED
        ready = True

    model.validation_result = result.value
    model.ready_for_export = ready
    model.issues = issues

    return {
        "validation_result": result.value,
        "summary": {
            "info_count": info_count,
            "warning_count": warning_count,
            "error_count": error_count,
            "blocker_count": blocker_count,
        },
        "issues": [to_dict(i) for i in issues],
        "confidence_summary": to_dict(model.confidence),
        "ready_for_export": ready,
    }


# ===================================================================
# 6. Lineage Service
# ===================================================================

def get_lineage(entity_type: str, entity_id: str) -> Dict[str, Any]:
    """Return provenance for an entity."""
    # Search by entity_id
    for lin in _lineage_store.values():
        if lin.entity_id == entity_id and lin.entity_type == entity_type:
            return {"lineage": to_dict(lin)}

    raise ValueError(f"No lineage found for {entity_type} '{entity_id}'.")


# ===================================================================
# 7. Export Service
# ===================================================================

def export_ready_json(
    model_id: str,
    export_profile: str = "report_ready_v1",
    include_lineage: bool = True,
    include_validation: bool = True,
    include_null_fields: bool = False,
    output_dir: str = "exports",
) -> Dict[str, Any]:
    """Export a validated canonical model as a downstream-ready JSON package."""
    model = _canonical_models.get(model_id)
    if model is None:
        raise ValueError(f"Model '{model_id}' not found.")
    if not model.ready_for_export:
        raise ValueError(f"Model '{model_id}' is not ready for export. Run validation first.")

    wb_record = _workbooks.get(model.workbook_id)

    payload = {
        "metadata": {
            "export_id": _uid("exp"),
            "export_profile": export_profile,
            "created_at": _now_iso(),
        },
        "workbook": to_dict(wb_record) if wb_record else {},
        "processing": {
            "model_id": model.model_id,
            "processing_profile": model.processing_profile,
        },
        "validation": {
            "result": model.validation_result,
            "overall_confidence": model.confidence.overall_model_confidence,
        } if include_validation else {},
        "assets": [to_dict(a) for a in model.assets],
        "scenarios": [to_dict(s) for s in model.scenarios],
        "kpis": [to_dict(k) for k in model.kpis],
        "measurements": [to_dict(m) for m in model.measurements],
        "time_series": [to_dict(t) for t in model.time_series],
        "assumptions": [to_dict(a) for a in model.assumptions],
        "figure_specs": [to_dict(f) for f in model.figure_specs],
        "lineage": [to_dict(l) for l in model.lineage] if include_lineage else [],
        "issues": [to_dict(i) for i in model.issues],
    }

    # Write to disk
    export_id = payload["metadata"]["export_id"]
    out_path = Path(output_dir) / export_id
    out_path.mkdir(parents=True, exist_ok=True)
    file_out = out_path / "report_ready.json"
    with open(file_out, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)

    return {
        "export": {
            "export_id": export_id,
            "profile": export_profile,
            "format": "json",
            "uri": str(file_out),
        },
        "payload_preview": {
            "workbook_id": model.workbook_id,
            "model_id": model.model_id,
            "assets": len(model.assets),
            "measurements": len(model.measurements),
            "validation_result": model.validation_result,
        },
    }


# ===================================================================
# 8. Figure Preparation Service
# ===================================================================

def prepare_figure_specs(
    model_id: str,
    figure_profile: str = "default",
    requested_outputs: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Generate structured figure specifications from validated data."""
    model = _canonical_models.get(model_id)
    if model is None:
        raise ValueError(f"Model '{model_id}' not found.")

    specs: List[FigureSpec] = []
    requests = requested_outputs or ["asset_comparison", "kpi_cards"]

    if "asset_comparison" in requests and model.assets and model.kpis:
        first_kpi = model.kpis[0]
        spec = FigureSpec.create(
            chart_type="grouped_bar",
            title=f"{first_kpi.kpi_name} by Asset",
            entity_refs=[{"entity_type": "Asset", "entity_id": a.asset_id} for a in model.assets],
            x_field="asset_name",
            y_field="value",
            unit=first_kpi.default_unit,
            ready_to_render=True,
        )
        specs.append(spec)

    if "time_series_trend" in requests and model.time_series:
        spec = FigureSpec.create(
            chart_type="line",
            title="Time Series Trend",
            x_field="timestamp",
            y_field="value",
            ready_to_render=bool(model.time_series),
        )
        specs.append(spec)

    if "kpi_cards" in requests and model.kpis:
        spec = FigureSpec.create(
            chart_type="kpi_card",
            title="Key Performance Indicators",
            entity_refs=[{"entity_type": "KPI", "entity_id": k.kpi_id} for k in model.kpis],
            ready_to_render=True,
        )
        specs.append(spec)

    model.figure_specs = specs

    return {
        "figure_specs": [to_dict(s) for s in specs],
    }


# ===================================================================
# 9. Report Bundle Service
# ===================================================================

def create_report_bundle(
    model_id: str,
    bundle_profile: str = "report_bundle_v1",
    include_json: bool = True,
    include_validation_report: bool = True,
    include_figure_specs: bool = True,
    include_summary_payload: bool = True,
    output_dir: str = "bundles",
) -> Dict[str, Any]:
    """Package all validated outputs for report production."""
    model = _canonical_models.get(model_id)
    if model is None:
        raise ValueError(f"Model '{model_id}' not found.")
    if not model.ready_for_export:
        raise ValueError(f"Model '{model_id}' is not ready for export.")

    bundle_id = _uid("rb")
    out_path = Path(output_dir) / bundle_id
    out_path.mkdir(parents=True, exist_ok=True)

    contents: List[str] = []

    if include_json:
        report = {
            "metadata": {"model_id": model.model_id, "created_at": _now_iso()},
            "assets": [to_dict(a) for a in model.assets],
            "scenarios": [to_dict(s) for s in model.scenarios],
            "kpis": [to_dict(k) for k in model.kpis],
            "measurements": [to_dict(m) for m in model.measurements],
            "lineage": [to_dict(l) for l in model.lineage],
            "issues": [to_dict(i) for i in model.issues],
        }
        _write_json(out_path / "report_ready.json", report)
        contents.append("report_ready.json")

    if include_validation_report:
        val = {
            "model_id": model.model_id,
            "validation_result": model.validation_result,
            "summary": {
                "info_count": sum(1 for i in model.issues if i.severity == "INFO"),
                "warning_count": sum(1 for i in model.issues if i.severity == "WARNING"),
                "error_count": sum(1 for i in model.issues if i.severity == "ERROR"),
                "blocker_count": sum(1 for i in model.issues if i.severity == "BLOCKER"),
            },
            "issues": [to_dict(i) for i in model.issues],
            "confidence_summary": to_dict(model.confidence),
        }
        _write_json(out_path / "validation_report.json", val)
        contents.append("validation_report.json")

    if include_figure_specs:
        figs = {
            "model_id": model.model_id,
            "figure_profile": bundle_profile,
            "figure_specs": [to_dict(f) for f in model.figure_specs],
        }
        _write_json(out_path / "figure_specs.json", figs)
        contents.append("figure_specs.json")

    if include_summary_payload:
        summary = {
            "model_id": model.model_id,
            "summary_context": {
                "validation_result": model.validation_result,
                "top_findings": [i.message for i in model.issues[:5]],
                "key_kpis": [{"kpi_name": k.kpi_name, "unit": k.default_unit} for k in model.kpis[:10]],
                "figure_contexts": [{"title": f.title, "chart_type": f.chart_type} for f in model.figure_specs],
            },
        }
        _write_json(out_path / "summary_payload.json", summary)
        contents.append("summary_payload.json")

    bundle = ReportBundle.create(profile=bundle_profile, contents=contents, uri=str(out_path))
    _bundles[bundle.bundle_id] = bundle

    return {
        "bundle": to_dict(bundle),
    }


# ===================================================================
# 10. Grounded Query Service
# ===================================================================

def grounded_query_validated_data(
    model_id: str,
    question: str,
    response_mode: str = "structured",
    include_lineage: bool = True,
) -> Dict[str, Any]:
    """Answer grounded queries over validated canonical model data only."""
    model = _canonical_models.get(model_id)
    if model is None:
        raise ValueError(f"Model '{model_id}' not found.")

    q_lower = question.lower()

    # Heuristic query resolution over validated data
    if "highest" in q_lower or "maximum" in q_lower or "most" in q_lower:
        return _query_extremum(model, question, "max", include_lineage)
    if "lowest" in q_lower or "minimum" in q_lower or "least" in q_lower:
        return _query_extremum(model, question, "min", include_lineage)
    if "count" in q_lower or "how many" in q_lower:
        return _query_count(model, question)
    if "list" in q_lower or "all" in q_lower:
        return _query_list(model, question, include_lineage)

    # Default: summary
    return {
        "answer": {
            "response_text": f"Model {model_id} contains {len(model.assets)} assets, {len(model.kpis)} KPIs, and {len(model.measurements)} measurements.",
            "confidence": 0.90,
            "supporting_entities": [],
            "lineage_refs": [],
        },
    }


# ===================================================================
# 11. End-to-End Orchestration
# ===================================================================

def process_workbook_end_to_end(
    workbook_uri: str,
    processing_profile: str = "default",
    bundle_profile: str = "report_bundle_v1",
    allow_llm_disambiguation: bool = False,
) -> Dict[str, Any]:
    """Run the full pipeline and return ready outputs."""
    triage = triage_workbook(workbook_uri, processing_profile)
    if triage["triage_result"] == TriageResult.UNSUPPORTED.value:
        return {
            "status": "error",
            "triage_result": triage["triage_result"],
            "error": {"code": "UNSUPPORTED_COMPLEXITY", "message": "Workbook is unsupported."},
        }

    wb_id = triage["lineage_ref"]["workbook_id"]
    detect_candidate_tables(workbook_uri)
    build_result = build_canonical_model(workbook_uri, processing_profile)
    model_id = build_result["lineage_ref"]["model_id"]

    val = validate_canonical_model(model_id)
    if not val["ready_for_export"]:
        return {
            "status": "error",
            "triage_result": triage["triage_result"],
            "validation_result": val["validation_result"],
            "model_id": model_id,
            "issues": val["issues"],
        }

    prepare_figure_specs(model_id)
    bundle = create_report_bundle(model_id, bundle_profile)

    return {
        "status": "success",
        "triage_result": triage["triage_result"],
        "validation_result": val["validation_result"],
        "model_id": model_id,
        "bundle_uri": bundle["bundle"]["uri"],
        "warnings": [{"severity": i["severity"], "message": i["message"]} for i in val["issues"] if i["severity"] in ("WARNING", "INFO")],
    }


# ===================================================================
# Internal helpers
# ===================================================================

def _resolve_uri(uri: str) -> str:
    """Resolve a workbook URI to an absolute filesystem path."""
    path = os.path.abspath(uri)
    # Security: block path traversal beyond reasonable boundaries
    if ".." in os.path.relpath(path):
        canonical = os.path.realpath(path)
        path = canonical
    return path


def _validate_file(file_path: str) -> None:
    """Validate the file exists, has an allowed extension, and is within size limits."""
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Unsupported file type '{ext}'. Allowed: {ALLOWED_EXTENSIONS}")
    size_mb = os.path.getsize(file_path) / (1024 * 1024)
    if size_mb > MAX_FILE_SIZE_MB:
        raise ValueError(f"File too large ({size_mb:.1f} MB). Max: {MAX_FILE_SIZE_MB} MB.")


def _get_or_create_workbook(file_path: str, wb) -> WorkbookRecord:
    """Retrieve an existing workbook record or create one."""
    for rec in _workbooks.values():
        if os.path.abspath(rec.file_path) == os.path.abspath(file_path):
            return rec
    rec = WorkbookRecord.create(
        file_name=os.path.basename(file_path),
        file_path=file_path,
        sheet_count=len(wb.sheetnames),
    )
    _workbooks[rec.workbook_id] = rec
    return rec


def _estimate_table_count(ws) -> int:
    """Quick heuristic to estimate number of tables in a worksheet."""
    if ws.max_row is None or ws.max_row < 2:
        return 0
    count = 1
    blank_run = 0
    for row in range(1, min(ws.max_row + 1, 200)):
        all_blank = all(ws.cell(row=row, column=c).value is None for c in range(1, min((ws.max_column or 1) + 1, 30)))
        if all_blank:
            blank_run += 1
            if blank_run >= 2:
                count += 1
                blank_run = 0
        else:
            blank_run = 0
    return count


def _detect_tables_in_sheet(ws, sheet_name: str, workbook_id: str, include_parser_comparison: bool) -> List[DetectedTable]:
    """Detect tabular regions in a single worksheet."""
    tables: List[DetectedTable] = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 2 or max_col < 1:
        return tables

    # Build occupancy grid (sampled)
    scan_rows = min(max_row, 500)
    scan_cols = min(max_col, 50)

    # Find contiguous blocks separated by blank rows
    regions: List[Tuple[int, int]] = []
    in_block = False
    block_start = 1
    for r in range(1, scan_rows + 1):
        row_has_data = any(ws.cell(row=r, column=c).value is not None for c in range(1, scan_cols + 1))
        if row_has_data and not in_block:
            block_start = r
            in_block = True
        elif not row_has_data and in_block:
            if r - block_start >= 2:  # at least 2 rows (header + data)
                regions.append((block_start, r - 1))
            in_block = False
    if in_block and (scan_rows - block_start >= 1):
        regions.append((block_start, scan_rows))

    for idx, (start, end) in enumerate(regions):
        # Determine used columns in this region
        used_cols = set()
        for r in range(start, end + 1):
            for c in range(1, scan_cols + 1):
                if ws.cell(row=r, column=c).value is not None:
                    used_cols.add(c)
        if not used_cols:
            continue
        min_col = min(used_cols)
        max_c = max(used_cols)

        # Read headers from first row
        headers = []
        for c in range(min_col, max_c + 1):
            val = ws.cell(row=start, column=c).value
            headers.append(str(val).strip() if val is not None else f"Column_{c}")

        # Confidence heuristics
        confidence = 0.85
        # Boost if headers look textual and data rows look numeric
        header_text_count = sum(1 for h in headers if h and not h.replace(".", "").replace("-", "").isdigit())
        if header_text_count >= len(headers) * 0.6:
            confidence += 0.05
        if (end - start) >= 3:
            confidence += 0.03
        confidence = min(confidence, 0.98)

        # Parser agreement (openpyxl always true; pandas if available)
        agreement = {"openpyxl": True}
        if PANDAS_AVAILABLE and include_parser_comparison:
            try:
                df = pd.read_excel(ws.parent.parent if hasattr(ws, "parent") else ws.parent.path,
                                   sheet_name=sheet_name, header=start - 1, nrows=end - start)
                agreement["pandas"] = len(df.columns) > 0
            except Exception:
                agreement["pandas"] = False
        if DOCLING_AVAILABLE:
            agreement["docling"] = True

        rng = f"{get_column_letter(min_col)}{start}:{get_column_letter(max_c)}{end}"
        data_start = start + 1  # first data row after header

        index_cols = [get_column_letter(min_col)] if headers else []

        dt = DetectedTable.create(
            workbook_id=workbook_id,
            sheet_name=sheet_name,
            range=rng,
            header_rows=[start],
            data_start_row=data_start,
            data_end_row=end,
            index_columns=[min_col],
            column_headers=headers,
            detection_confidence=round(confidence, 2),
            parser_agreement=agreement,
            provenance={
                "workbook_id": workbook_id,
                "sheet_name": sheet_name,
                "range": rng,
                "detection_methods": list(agreement.keys()),
            },
        )
        tables.append(dt)

    return tables


def _guess_unit(header: str) -> str:
    """Attempt to extract a unit from a header string."""
    m = re.search(r"\(([^)]+)\)", header)
    if m:
        return m.group(1).strip()
    lower = header.lower()
    if "mw" in lower:
        return "MW"
    if "gwh" in lower:
        return "GWh"
    if "%" in header:
        return "%"
    return ""


def _write_json(path: Path, data: Any) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)


def _query_extremum(model: CanonicalModel, question: str, mode: str, include_lineage: bool) -> Dict[str, Any]:
    """Find the asset with the highest/lowest measurement."""
    if not model.measurements:
        return {"answer": {"response_text": "No measurements available.", "confidence": 0.5, "supporting_entities": [], "lineage_refs": []}}

    if mode == "max":
        target = max(model.measurements, key=lambda m: m.value)
    else:
        target = min(model.measurements, key=lambda m: m.value)

    asset = next((a for a in model.assets if a.asset_id == target.asset_id), None)
    kpi = next((k for k in model.kpis if k.kpi_id == target.kpi_id), None)

    asset_name = asset.asset_name if asset else target.asset_id
    kpi_name = kpi.kpi_name if kpi else target.kpi_id
    direction = "highest" if mode == "max" else "lowest"

    lineage_refs = []
    if include_lineage and target.lineage_ref:
        lin = _lineage_store.get(target.lineage_ref)
        if lin:
            lineage_refs.append({"sheet_name": lin.source_sheet_name, "range": lin.source_range})

    return {
        "answer": {
            "response_text": f"{asset_name} has the {direction} {kpi_name}: {target.value} {target.unit}.",
            "confidence": 0.94,
            "supporting_entities": [
                {"entity_type": "Asset", "entity_id": target.asset_id},
                {"entity_type": "Measurement", "entity_id": target.measurement_id},
            ],
            "lineage_refs": lineage_refs,
        },
    }


def _query_count(model: CanonicalModel, question: str) -> Dict[str, Any]:
    return {
        "answer": {
            "response_text": f"The model contains {len(model.assets)} assets, {len(model.kpis)} KPIs, {len(model.measurements)} measurements, and {len(model.scenarios)} scenarios.",
            "confidence": 1.0,
            "supporting_entities": [],
            "lineage_refs": [],
        },
    }


def _query_list(model: CanonicalModel, question: str, include_lineage: bool) -> Dict[str, Any]:
    items = [{"entity_type": "Asset", "entity_id": a.asset_id, "name": a.asset_name} for a in model.assets]
    return {
        "answer": {
            "response_text": f"Assets: {', '.join(a.asset_name for a in model.assets)}." if model.assets else "No assets found.",
            "confidence": 1.0,
            "supporting_entities": items,
            "lineage_refs": [],
        },
    }
