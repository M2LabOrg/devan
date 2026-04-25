"""
Generate comprehensive test Excel files covering different scenarios.

Test Scenarios:
1. Simple single-table spreadsheet
2. Multi-table with gaps
3. Formulas and calculations
4. Merged cells and formatting
5. Charts and figures
6. Empty rows/columns
7. Data validation and dropdowns
8. Pivot table-like structure
9. Financial statements
10. Scientific data with units
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import random


def create_test_1_simple_table():
    """Test 1: Simple single-table spreadsheet"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    # Simple employee table
    headers = ["Employee ID", "Name", "Department", "Salary", "Hire Date"]
    ws.append(headers)
    
    # Make headers bold
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # Add data
    employees = [
        ["E001", "Alice Johnson", "Engineering", 95000, datetime(2020, 3, 15)],
        ["E002", "Bob Smith", "Sales", 75000, datetime(2019, 7, 22)],
        ["E003", "Carol White", "Engineering", 98000, datetime(2021, 1, 10)],
        ["E004", "David Brown", "Marketing", 72000, datetime(2020, 11, 5)],
        ["E005", "Eve Davis", "HR", 68000, datetime(2018, 9, 30)],
    ]
    
    for emp in employees:
        ws.append(emp)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    
    path = "excel_files/test_1_simple_table.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def create_test_2_multi_table_gaps():
    """Test 2: Multiple tables with gaps and instructions"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Title
    ws['A1'] = "Q1 2024 Sales Report"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Instruction
    ws['A2'] = "Note: All amounts in USD. Updated monthly."
    ws['A2'].font = Font(italic=True)
    
    # Table 1: Sales by Region (starts at row 4)
    ws['A4'] = "Sales by Region"
    ws['A4'].font = Font(bold=True)
    
    headers1 = ["Region", "January", "February", "March", "Total"]
    for i, h in enumerate(headers1, 1):
        ws.cell(row=5, column=i, value=h).font = Font(bold=True)
    
    sales_data = [
        ["North", 45000, 48000, 52000, 145000],
        ["South", 38000, 41000, 39000, 118000],
        ["East", 52000, 55000, 58000, 165000],
        ["West", 41000, 43000, 46000, 130000],
    ]
    
    for r, row_data in enumerate(sales_data, 6):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    
    # Gap (rows 10-12 empty)
    
    # Table 2: Top Products (starts at row 13)
    ws['A13'] = "Top 5 Products"
    ws['A13'].font = Font(bold=True)
    
    headers2 = ["Product", "Units Sold", "Revenue"]
    for i, h in enumerate(headers2, 1):
        ws.cell(row=14, column=i, value=h).font = Font(bold=True)
    
    products = [
        ["Widget A", 1200, 36000],
        ["Widget B", 980, 29400],
        ["Gadget X", 850, 42500],
        ["Gadget Y", 720, 28800],
        ["Tool Z", 650, 19500],
    ]
    
    for r, row_data in enumerate(products, 15):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    
    path = "excel_files/test_2_multi_table_gaps.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def create_test_3_formulas():
    """Test 3: Spreadsheet with formulas and calculations"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"
    
    ws['A1'] = "Department Budget 2024"
    ws['A1'].font = Font(bold=True)
    
    headers = ["Department", "Q1", "Q2", "Q3", "Q4", "Total", "% of Budget"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h).font = Font(bold=True)
    
    departments = [
        ["Engineering", 50000, 52000, 48000, 55000],
        ["Sales", 35000, 38000, 40000, 42000],
        ["Marketing", 25000, 28000, 30000, 32000],
        ["Operations", 40000, 41000, 39000, 43000],
    ]
    
    for r, dept_data in enumerate(departments, 3):
        ws.cell(row=r, column=1, value=dept_data[0])
        for c, val in enumerate(dept_data[1:], 2):
            ws.cell(row=r, column=c, value=val)
        
        # Formula for Total (column F)
        ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
    
    # Grand total row
    ws.cell(row=7, column=1, value="Grand Total").font = Font(bold=True)
    for c in range(2, 7):
        col_letter = get_column_letter(c)
        ws.cell(row=7, column=c, value=f"=SUM({col_letter}3:{col_letter}6)")
    
    # % of Budget formulas (column G)
    for r in range(3, 7):
        ws.cell(row=r, column=7, value=f"=F{r}/$F$7")
        ws.cell(row=r, column=7).number_format = '0.0%'
    
    path = "excel_files/test_3_formulas.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def create_test_4_merged_cells():
    """Test 4: Merged cells and complex formatting"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quarterly Report"
    
    # Merged title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Quarterly Performance Report - 2024"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Merged subtitle
    ws.merge_cells('A2:F2')
    ws['A2'] = "Confidential - Internal Use Only"
    ws['A2'].font = Font(italic=True, color="CC0000")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers with merged cells for quarters
    ws.merge_cells('B4:C4')
    ws['B4'] = "Q1"
    ws['B4'].font = Font(bold=True)
    ws['B4'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('D4:E4')
    ws['D4'] = "Q2"
    ws['D4'].font = Font(bold=True)
    ws['D4'].alignment = Alignment(horizontal='center')
    
    # Sub-headers
    ws['A5'] = "Metric"
    ws['B5'] = "Target"
    ws['C5'] = "Actual"
    ws['D5'] = "Target"
    ws['E5'] = "Actual"
    ws['F5'] = "YTD Total"
    
    for col in range(1, 7):
        ws.cell(row=5, column=col).font = Font(bold=True)
    
    # Data
    metrics = [
        ["Revenue", 100000, 105000, 110000, 108000, 213000],
        ["Profit", 25000, 28000, 30000, 29000, 57000],
        ["Customers", 500, 520, 550, 540, 1060],
    ]
    
    for r, row_data in enumerate(metrics, 6):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    
    path = "excel_files/test_4_merged_cells.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def create_test_5_with_charts():
    """Test 5: Spreadsheet with embedded charts"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Trends"
    
    ws['A1'] = "Monthly Sales Trends"
    ws['A1'].font = Font(bold=True)
    
    # Data table
    headers = ["Month", "Product A", "Product B", "Product C"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h).font = Font(bold=True)
    
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    for i, month in enumerate(months, 3):
        ws.cell(row=i, column=1, value=month)
        ws.cell(row=i, column=2, value=random.randint(80, 120))
        ws.cell(row=i, column=3, value=random.randint(60, 100))
        ws.cell(row=i, column=4, value=random.randint(70, 110))
    
    # Add a line chart
    chart = LineChart()
    chart.title = "Product Sales Over Time"
    chart.style = 10
    chart.y_axis.title = "Sales"
    chart.x_axis.title = "Month"
    
    data = Reference(ws, min_col=2, min_row=2, max_col=4, max_row=8)
    cats = Reference(ws, min_col=1, min_row=3, max_row=8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "F2")
    
    # Add a second table below
    ws['A12'] = "Product Summary"
    ws['A12'].font = Font(bold=True)
    
    summary_headers = ["Product", "Total Sales", "Avg Monthly"]
    for i, h in enumerate(summary_headers, 1):
        ws.cell(row=13, column=i, value=h).font = Font(bold=True)
    
    ws.cell(row=14, column=1, value="Product A")
    ws.cell(row=14, column=2, value="=SUM(B3:B8)")
    ws.cell(row=14, column=3, value="=AVERAGE(B3:B8)")
    
    ws.cell(row=15, column=1, value="Product B")
    ws.cell(row=15, column=2, value="=SUM(C3:C8)")
    ws.cell(row=15, column=3, value="=AVERAGE(C3:C8)")
    
    ws.cell(row=16, column=1, value="Product C")
    ws.cell(row=16, column=2, value="=SUM(D3:D8)")
    ws.cell(row=16, column=3, value="=AVERAGE(D3:D8)")
    
    path = "excel_files/test_5_with_charts.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def create_test_6_empty_rows_cols():
    """Test 6: Tables with empty rows and columns in between"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sparse Data"
    
    # Table 1 in columns A-C
    ws['A1'] = "Team A"
    ws['A1'].font = Font(bold=True)
    
    ws['A2'] = "Name"
    ws['B2'] = "Score"
    ws.cell(row=2, column=1).font = Font(bold=True)
    ws.cell(row=2, column=2).font = Font(bold=True)
    
    team_a = [["Alice", 95], ["Bob", 87], ["Carol", 92]]
    for r, row_data in enumerate(team_a, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    
    # Empty column D
    
    # Table 2 in columns E-F
    ws['E1'] = "Team B"
    ws['E1'].font = Font(bold=True)
    
    ws['E2'] = "Name"
    ws['F2'] = "Score"
    ws.cell(row=2, column=5).font = Font(bold=True)
    ws.cell(row=2, column=6).font = Font(bold=True)
    
    team_b = [["David", 89], ["Eve", 94], ["Frank", 88]]
    for r, row_data in enumerate(team_b, 3):
        for c, val in enumerate(row_data, 5):
            ws.cell(row=r, column=c, value=val)
    
    # Empty rows 7-8
    
    # Table 3 spanning columns A-F at row 10
    ws['A10'] = "Combined Results"
    ws['A10'].font = Font(bold=True)
    
    combined_headers = ["Team", "Member", "Score", "Grade"]
    for i, h in enumerate(combined_headers, 1):
        ws.cell(row=11, column=i, value=h).font = Font(bold=True)
    
    combined = [
        ["Team A", "Alice", 95, "A"],
        ["Team A", "Bob", 87, "B"],
        ["Team B", "David", 89, "B"],
        ["Team B", "Eve", 94, "A"],
    ]
    
    for r, row_data in enumerate(combined, 12):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    
    path = "excel_files/test_6_empty_rows_cols.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def create_test_7_financial_statement():
    """Test 7: Financial statement with hierarchical structure"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "Income Statement - FY 2024"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Line Item", "Q1", "Q2", "YTD"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h).font = Font(bold=True)
    
    # Revenue section
    ws['A4'] = "Revenue"
    ws['A4'].font = Font(bold=True)
    
    revenue_items = [
        ["  Product Sales", 500000, 520000, 1020000],
        ["  Service Revenue", 150000, 160000, 310000],
        ["  Other Income", 20000, 25000, 45000],
    ]
    
    for r, row_data in enumerate(revenue_items, 5):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    
    ws['A8'] = "Total Revenue"
    ws['A8'].font = Font(bold=True)
    for c in range(2, 5):
        ws.cell(row=8, column=c, value=f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}7)")
        ws.cell(row=8, column=c).font = Font(bold=True)
    
    # Expenses section
    ws['A10'] = "Expenses"
    ws['A10'].font = Font(bold=True)
    
    expense_items = [
        ["  Cost of Goods Sold", 300000, 310000, 610000],
        ["  Operating Expenses", 150000, 155000, 305000],
        ["  Marketing", 50000, 55000, 105000],
        ["  R&D", 40000, 42000, 82000],
    ]
    
    for r, row_data in enumerate(expense_items, 11):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    
    ws['A15'] = "Total Expenses"
    ws['A15'].font = Font(bold=True)
    for c in range(2, 5):
        ws.cell(row=15, column=c, value=f"=SUM({get_column_letter(c)}11:{get_column_letter(c)}14)")
        ws.cell(row=15, column=c).font = Font(bold=True)
    
    # Net Income
    ws['A17'] = "Net Income"
    ws['A17'].font = Font(bold=True, color="006100")
    for c in range(2, 5):
        ws.cell(row=17, column=c, value=f"={get_column_letter(c)}8-{get_column_letter(c)}15")
        ws.cell(row=17, column=c).font = Font(bold=True, color="006100")
    
    path = "excel_files/test_7_financial_statement.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def create_test_8_scientific_data():
    """Test 8: Scientific data with units and measurements"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lab Results"
    
    ws['A1'] = "Laboratory Test Results - Batch #2024-03"
    ws['A1'].font = Font(bold=True)
    
    ws['A2'] = "Note: All measurements taken at 25°C, 1 atm pressure"
    ws['A2'].font = Font(italic=True)
    
    # Table 1: Chemical Analysis
    ws['A4'] = "Chemical Composition Analysis"
    ws['A4'].font = Font(bold=True)
    
    headers1 = ["Sample ID", "pH", "Concentration (mg/L)", "Temperature (°C)", "Status"]
    for i, h in enumerate(headers1, 1):
        ws.cell(row=5, column=i, value=h).font = Font(bold=True)
    
    samples = [
        ["S001", 7.2, 45.3, 25.1, "Pass"],
        ["S002", 6.8, 52.7, 24.9, "Pass"],
        ["S003", 7.5, 48.1, 25.3, "Pass"],
        ["S004", 8.1, 61.2, 25.0, "Retest"],
        ["S005", 7.0, 43.8, 25.2, "Pass"],
    ]
    
    for r, row_data in enumerate(samples, 6):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    
    # Gap
    
    # Table 2: Statistical Summary
    ws['A13'] = "Statistical Summary"
    ws['A13'].font = Font(bold=True)
    
    headers2 = ["Metric", "pH", "Concentration (mg/L)", "Temperature (°C)"]
    for i, h in enumerate(headers2, 1):
        ws.cell(row=14, column=i, value=h).font = Font(bold=True)
    
    stats = [
        ["Mean", "=AVERAGE(B6:B10)", "=AVERAGE(C6:C10)", "=AVERAGE(D6:D10)"],
        ["Std Dev", "=STDEV(B6:B10)", "=STDEV(C6:C10)", "=STDEV(D6:D10)"],
        ["Min", "=MIN(B6:B10)", "=MIN(C6:C10)", "=MIN(D6:D10)"],
        ["Max", "=MAX(B6:B10)", "=MAX(C6:C10)", "=MAX(D6:D10)"],
    ]
    
    for r, row_data in enumerate(stats, 15):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    
    path = "excel_files/test_8_scientific_data.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def create_test_9_pivot_like():
    """Test 9: Pivot table-like structure with cross-tabulation"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Matrix"
    
    ws['A1'] = "Sales by Product and Region - 2024"
    ws['A1'].font = Font(bold=True)
    
    # Cross-tab table
    ws['A3'] = "Product \\ Region"
    ws['A3'].font = Font(bold=True)
    
    regions = ["North", "South", "East", "West", "Total"]
    for i, region in enumerate(regions, 2):
        ws.cell(row=3, column=i, value=region).font = Font(bold=True)
    
    products = ["Widget A", "Widget B", "Gadget X", "Gadget Y", "Tool Z"]
    
    for i, product in enumerate(products, 4):
        ws.cell(row=i, column=1, value=product).font = Font(bold=True)
        
        # Random sales data
        for j in range(2, 6):
            ws.cell(row=i, column=j, value=random.randint(10000, 50000))
        
        # Total column
        ws.cell(row=i, column=6, value=f"=SUM(B{i}:E{i})")
    
    # Total row
    ws.cell(row=9, column=1, value="Total").font = Font(bold=True)
    for c in range(2, 7):
        col_letter = get_column_letter(c)
        ws.cell(row=9, column=c, value=f"=SUM({col_letter}4:{col_letter}8)")
        ws.cell(row=9, column=c).font = Font(bold=True)
    
    path = "excel_files/test_9_pivot_like.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def create_test_10_mixed_data_types():
    """Test 10: Mixed data types and edge cases"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed Data"
    
    ws['A1'] = "Data Type Testing"
    ws['A1'].font = Font(bold=True)
    
    headers = ["Type", "Value", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h).font = Font(bold=True)
    
    test_data = [
        ["Integer", 42, "Simple integer"],
        ["Float", 3.14159, "Decimal number"],
        ["Currency", 1234.56, "Dollar amount"],
        ["Percentage", 0.75, "75%"],
        ["Date", datetime(2024, 3, 15), "Date value"],
        ["Time", datetime(2024, 1, 1, 14, 30), "Time value"],
        ["Boolean", True, "True/False"],
        ["Text", "Hello World", "String value"],
        ["Formula", "=2+2", "Calculation"],
        ["Empty", None, "Null value"],
        ["Large Number", 1234567890, "Big integer"],
        ["Negative", -500, "Negative value"],
    ]
    
    for r, row_data in enumerate(test_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            
            # Format specific cells
            if row_data[0] == "Currency":
                ws.cell(row=r, column=2).number_format = '$#,##0.00'
            elif row_data[0] == "Percentage":
                ws.cell(row=r, column=2).number_format = '0%'
            elif row_data[0] == "Date":
                ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'
            elif row_data[0] == "Time":
                ws.cell(row=r, column=2).number_format = 'hh:mm:ss'
    
    path = "excel_files/test_10_mixed_data_types.xlsx"
    wb.save(path)
    print(f"✓ Created: {path}")
    return path


def main():
    """Generate all test files"""
    print("Generating comprehensive test Excel files...\n")
    
    os.makedirs("excel_files", exist_ok=True)
    
    test_files = []
    
    test_files.append(create_test_1_simple_table())
    test_files.append(create_test_2_multi_table_gaps())
    test_files.append(create_test_3_formulas())
    test_files.append(create_test_4_merged_cells())
    test_files.append(create_test_5_with_charts())
    test_files.append(create_test_6_empty_rows_cols())
    test_files.append(create_test_7_financial_statement())
    test_files.append(create_test_8_scientific_data())
    test_files.append(create_test_9_pivot_like())
    test_files.append(create_test_10_mixed_data_types())
    
    print(f"\n✅ Successfully generated {len(test_files)} test files")
    return test_files


if __name__ == "__main__":
    main()
