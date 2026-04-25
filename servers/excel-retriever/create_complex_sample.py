"""
Create a complex sample Excel file to test the intelligent extractor.

This generates a realistic Excel file with:
- Multiple tabs
- Multiple tables per tab
- Instruction text between tables
- Merged cells
- Charts/figures
- Named ranges
- Mixed data types
- Empty rows/columns separating content
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import os
import random
from datetime import datetime, timedelta

def create_complex_excel():
    wb = openpyxl.Workbook()

    # ─── Sheet 1: Dashboard / Summary ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Dashboard"

    # Title (merged cells)
    ws1.merge_cells('A1:F1')
    ws1['A1'] = "Q4 2025 Performance Dashboard"
    ws1['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws1['A1'].alignment = Alignment(horizontal='center')

    # Instruction text
    ws1['A3'] = "Note: All figures are in thousands (USD). Updated quarterly."
    ws1['A3'].font = Font(italic=True, color="666666")

    ws1['A4'] = "Warning: Revenue targets for APAC region are provisional estimates."
    ws1['A4'].font = Font(italic=True, color="CC0000")

    # ── Table 1: KPI Summary (starts row 6) ──
    ws1['A6'] = "Key Performance Indicators"
    ws1['A6'].font = Font(size=12, bold=True)

    kpi_headers = ["KPI", "Target", "Actual", "Variance", "Status"]
    for i, h in enumerate(kpi_headers, 1):
        cell = ws1.cell(row=7, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    kpi_data = [
        ["Revenue", 15000, 14200, -800, "Behind"],
        ["Gross Margin", 0.42, 0.45, 0.03, "On Track"],
        ["Customer Satisfaction", 4.5, 4.7, 0.2, "Exceeded"],
        ["Employee Retention", 0.92, 0.89, -0.03, "At Risk"],
        ["New Contracts", 25, 31, 6, "Exceeded"],
        ["Safety Incidents", 0, 1, 1, "Behind"],
    ]

    for r, row_data in enumerate(kpi_data, 8):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Empty rows gap

    # ── Table 2: Regional Breakdown (starts row 16) ──
    ws1['A16'] = "Regional Revenue Breakdown"
    ws1['A16'].font = Font(size=12, bold=True)

    region_headers = ["Region", "Q1", "Q2", "Q3", "Q4", "Total"]
    for i, h in enumerate(region_headers, 1):
        cell = ws1.cell(row=17, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    regions = [
        ["EMEA", 3200, 3400, 3600, 3800, 14000],
        ["Americas", 4100, 4300, 4500, 4200, 17100],
        ["APAC", 1800, 2100, 2300, 2500, 8700],
        ["Middle East", 900, 1100, 1200, 1400, 4600],
    ]

    for r, row_data in enumerate(regions, 18):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Add a chart
    chart1 = BarChart()
    chart1.title = "Regional Revenue by Quarter"
    chart1.type = "col"
    chart1.style = 10
    data = Reference(ws1, min_col=2, min_row=17, max_col=5, max_row=21)
    cats = Reference(ws1, min_col=1, min_row=18, max_row=21)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    ws1.add_chart(chart1, "H6")

    # ── Table 3: Bottom of sheet, different topic ──
    ws1['A25'] = "Important: The following cost centers require management review."
    ws1['A25'].font = Font(bold=True, color="CC0000")

    cost_headers = ["Cost Center", "Budget", "Spent", "Remaining", "% Used"]
    for i, h in enumerate(cost_headers, 1):
        cell = ws1.cell(row=27, column=i, value=h)
        cell.font = Font(bold=True)

    costs = [
        ["IT Infrastructure", 500, 480, 20, 0.96],
        ["Marketing", 300, 310, -10, 1.03],
        ["R&D", 800, 650, 150, 0.81],
        ["Operations", 1200, 1180, 20, 0.98],
    ]

    for r, row_data in enumerate(costs, 28):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # ─── Sheet 2: Project Details ───────────────────────────────────────────
    ws2 = wb.create_sheet("Projects")

    ws2.merge_cells('A1:H1')
    ws2['A1'] = "Active Project Portfolio"
    ws2['A1'].font = Font(size=14, bold=True)

    ws2['A2'] = "Last updated: 2025-12-15. Contact: project.office@example.com"
    ws2['A2'].font = Font(italic=True)

    # ── Table 1: Active Projects ──
    ws2['A4'] = "Active Projects"
    ws2['A4'].font = Font(size=11, bold=True)

    proj_headers = ["Project ID", "Name", "Client", "Start Date", "End Date",
                    "Budget (K)", "Status", "Risk Level"]
    for i, h in enumerate(proj_headers, 1):
        cell = ws2.cell(row=5, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")

    projects = [
        ["PRJ-001", "North Sea Platform Inspection", "Equinor", datetime(2025, 1, 15), datetime(2025, 12, 31), 2400, "Active", "Medium"],
        ["PRJ-002", "Wind Farm Certification", "Ørsted", datetime(2025, 3, 1), datetime(2026, 6, 30), 1800, "Active", "Low"],
        ["PRJ-003", "Pipeline Integrity Assessment", "Shell", datetime(2025, 6, 1), datetime(2025, 11, 30), 950, "Completed", "Low"],
        ["PRJ-004", "LNG Terminal Safety Review", "TotalEnergies", datetime(2025, 2, 1), datetime(2025, 9, 30), 1100, "Active", "High"],
        ["PRJ-005", "Digital Twin Implementation", "BP", datetime(2025, 4, 15), datetime(2026, 4, 15), 3200, "Active", "Medium"],
        ["PRJ-006", "Subsea Cable Survey", "Vattenfall", datetime(2025, 7, 1), datetime(2026, 1, 31), 750, "Active", "Low"],
        ["PRJ-007", "FPSO Classification", "Petrobras", datetime(2025, 5, 1), datetime(2026, 8, 31), 4500, "Active", "High"],
        ["PRJ-008", "Hydrogen Storage Feasibility", "Air Liquide", datetime(2025, 8, 1), datetime(2026, 3, 31), 680, "Active", "Medium"],
    ]

    for r, row_data in enumerate(projects, 6):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # Gap and instructions
    ws2['A16'] = "Note: Projects marked High risk require weekly status updates to the Board."
    ws2['A16'].font = Font(italic=True, color="CC0000")

    ws2['A17'] = "See the Risk Analysis tab for detailed risk assessments."
    ws2['A17'].font = Font(italic=True)

    # ── Table 2: Milestones (different columns, same sheet) ──
    ws2['A19'] = "Upcoming Milestones"
    ws2['A19'].font = Font(size=11, bold=True)

    mile_headers = ["Project ID", "Milestone", "Due Date", "Owner", "Complete"]
    for i, h in enumerate(mile_headers, 1):
        cell = ws2.cell(row=20, column=i, value=h)
        cell.font = Font(bold=True)

    milestones = [
        ["PRJ-001", "Final Inspection Report", datetime(2025, 12, 15), "J. Hansen", "No"],
        ["PRJ-002", "Design Review Gate 3", datetime(2025, 11, 30), "M. Nielsen", "Yes"],
        ["PRJ-004", "Safety Case Submission", datetime(2025, 9, 15), "A. Dupont", "No"],
        ["PRJ-005", "Phase 1 Go-Live", datetime(2025, 12, 1), "S. Patel", "No"],
        ["PRJ-007", "Hull Structural Review", datetime(2026, 2, 28), "R. Silva", "No"],
    ]

    for r, row_data in enumerate(milestones, 21):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # ─── Sheet 3: Risk Analysis ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Risk Analysis")

    ws3.merge_cells('A1:F1')
    ws3['A1'] = "Risk Assessment Matrix"
    ws3['A1'].font = Font(size=14, bold=True)

    ws3['A3'] = "Instructions: Rate each risk on a 1-5 scale for both likelihood and impact."
    ws3['A3'].font = Font(italic=True)
    ws3['A4'] = "Risk Score = Likelihood × Impact. Scores >= 15 require mitigation plans."
    ws3['A4'].font = Font(italic=True)

    risk_headers = ["Risk ID", "Description", "Category", "Likelihood (1-5)",
                    "Impact (1-5)", "Risk Score", "Mitigation"]
    for i, h in enumerate(risk_headers, 1):
        cell = ws3.cell(row=6, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    risks = [
        ["R-001", "Supply chain delays for inspection equipment", "Operational", 3, 4, 12, "Establish backup suppliers"],
        ["R-002", "Regulatory changes in EU offshore wind", "Regulatory", 4, 5, 20, "Dedicated regulatory tracking team"],
        ["R-003", "Key personnel attrition", "HR", 3, 3, 9, "Succession planning program"],
        ["R-004", "Cybersecurity breach in digital twin platform", "Technology", 2, 5, 10, "Quarterly pen testing, SOC monitoring"],
        ["R-005", "Weather delays in North Sea operations", "Environmental", 4, 3, 12, "Buffer time in project schedules"],
        ["R-006", "Client payment defaults", "Financial", 2, 4, 8, "Credit insurance, milestone billing"],
        ["R-007", "FPSO structural fatigue findings", "Technical", 3, 5, 15, "Enhanced monitoring program"],
    ]

    for r, row_data in enumerate(risks, 7):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    # Gap, then a second table on the same sheet
    ws3['A17'] = "Risk Category Summary"
    ws3['A17'].font = Font(size=11, bold=True)

    cat_headers = ["Category", "Count", "Avg Score", "Max Score"]
    for i, h in enumerate(cat_headers, 1):
        cell = ws3.cell(row=18, column=i, value=h)
        cell.font = Font(bold=True)

    categories = [
        ["Operational", 2, 12, 12],
        ["Regulatory", 1, 20, 20],
        ["HR", 1, 9, 9],
        ["Technology", 1, 10, 10],
        ["Financial", 1, 8, 8],
        ["Technical", 1, 15, 15],
    ]

    for r, row_data in enumerate(categories, 19):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    # Add a pie chart
    pie = PieChart()
    pie.title = "Risks by Category"
    labels = Reference(ws3, min_col=1, min_row=19, max_row=24)
    data = Reference(ws3, min_col=2, min_row=18, max_row=24)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws3.add_chart(pie, "F17")

    # ─── Sheet 4: Instructions (text-heavy) ─────────────────────────────────
    ws4 = wb.create_sheet("Instructions")

    ws4.merge_cells('A1:D1')
    ws4['A1'] = "Data Entry Instructions"
    ws4['A1'].font = Font(size=14, bold=True)

    instructions = [
        (3, "1. General Guidelines"),
        (4, "All data must be entered in the designated cells only. Do not modify column headers or sheet structure."),
        (5, "Use the dropdown menus where available for consistent data entry."),
        (6, "Dates should be in YYYY-MM-DD format. Currency values in thousands (K)."),
        (8, "2. Project Data Entry"),
        (9, "Navigate to the 'Projects' tab to enter new project information."),
        (10, "Required fields: Project ID, Name, Client, Start Date, Budget, Status."),
        (11, "Project IDs follow the format PRJ-XXX where XXX is a sequential number."),
        (13, "3. Risk Assessment"),
        (14, "Use the 'Risk Analysis' tab for all risk-related entries."),
        (15, "Each risk must have a unique Risk ID (format: R-XXX)."),
        (16, "Risk scores are calculated automatically (Likelihood × Impact)."),
        (17, "Any risk with score >= 15 must have a documented mitigation plan."),
        (19, "4. Contacts"),
        (20, "For questions about this workbook, contact:"),
        (21, "  - Technical: engineering@example.com"),
        (22, "  - Process: project.office@example.com"),
        (23, "  - IT Support: helpdesk@example.com"),
    ]

    for row, text in instructions:
        ws4.cell(row=row, column=1, value=text)
        if text.startswith(("1.", "2.", "3.", "4.")):
            ws4.cell(row=row, column=1).font = Font(bold=True, size=11)

    # ── Small lookup table at the bottom ──
    ws4['A26'] = "Status Code Reference"
    ws4['A26'].font = Font(bold=True)

    ref_headers = ["Code", "Description", "Color"]
    for i, h in enumerate(ref_headers, 1):
        ws4.cell(row=27, column=i, value=h).font = Font(bold=True)

    ref_data = [
        ["Active", "Project is currently in execution", "Green"],
        ["On Hold", "Project is temporarily paused", "Yellow"],
        ["Completed", "Project has been delivered", "Blue"],
        ["Cancelled", "Project has been terminated", "Red"],
        ["At Risk", "Project has significant issues", "Orange"],
    ]

    for r, row_data in enumerate(ref_data, 28):
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)

    # ─── Sheet 5: Financial Details (multiple side-by-side tables) ──────────
    ws5 = wb.create_sheet("Financials")

    ws5.merge_cells('A1:L1')
    ws5['A1'] = "Financial Summary - Q4 2025"
    ws5['A1'].font = Font(size=14, bold=True)

    ws5['A2'] = "Confidential - Internal Use Only"
    ws5['A2'].font = Font(italic=True, color="CC0000")

    # ── Left table: Income Statement ──
    ws5['A4'] = "Income Statement"
    ws5['A4'].font = Font(bold=True, size=11)

    inc_headers = ["Item", "Q3 Actual", "Q4 Actual", "Q4 Budget"]
    for i, h in enumerate(inc_headers, 1):
        ws5.cell(row=5, column=i, value=h).font = Font(bold=True)

    income = [
        ["Revenue", 12500, 14200, 15000],
        ["COGS", -7200, -7800, -8100],
        ["Gross Profit", 5300, 6400, 6900],
        ["Operating Expenses", -3800, -4100, -4200],
        ["EBITDA", 1500, 2300, 2700],
        ["Depreciation", -400, -420, -400],
        ["EBIT", 1100, 1880, 2300],
        ["Interest", -200, -210, -200],
        ["Net Income", 900, 1670, 2100],
    ]

    for r, row_data in enumerate(income, 6):
        for c, val in enumerate(row_data, 1):
            ws5.cell(row=r, column=c, value=val)

    # ── Right table: Balance Sheet (side by side, starting col G) ──
    ws5['G4'] = "Balance Sheet Snapshot"
    ws5['G4'].font = Font(bold=True, size=11)

    bal_headers = ["Item", "Amount", "% of Total"]
    for i, h in enumerate(bal_headers, 7):
        ws5.cell(row=5, column=i, value=h).font = Font(bold=True)

    balance = [
        ["Cash & Equivalents", 8500, 0.22],
        ["Accounts Receivable", 6200, 0.16],
        ["Fixed Assets", 15000, 0.39],
        ["Intangible Assets", 4500, 0.12],
        ["Other Assets", 4300, 0.11],
        ["Total Assets", 38500, 1.00],
    ]

    for r, row_data in enumerate(balance, 6):
        for c, val in enumerate(row_data, 7):
            ws5.cell(row=r, column=c, value=val)

    # ── Below both: Cash Flow table ──
    ws5['A18'] = "Cash Flow Highlights"
    ws5['A18'].font = Font(bold=True, size=11)

    cf_headers = ["Category", "Q1", "Q2", "Q3", "Q4", "Annual"]
    for i, h in enumerate(cf_headers, 1):
        ws5.cell(row=19, column=i, value=h).font = Font(bold=True)

    cashflow = [
        ["Operating", 1200, 1400, 1500, 2300, 6400],
        ["Investing", -800, -600, -700, -900, -3000],
        ["Financing", -300, -300, -300, -300, -1200],
        ["Net Cash Flow", 100, 500, 500, 1100, 2200],
    ]

    for r, row_data in enumerate(cashflow, 20):
        for c, val in enumerate(row_data, 1):
            ws5.cell(row=r, column=c, value=val)

    # Add line chart for cash flow
    line = LineChart()
    line.title = "Quarterly Cash Flow Trend"
    data = Reference(ws5, min_col=2, min_row=19, max_col=5, max_row=23)
    cats = Reference(ws5, min_col=1, min_row=20, max_row=23)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws5.add_chart(line, "G18")

    # ─── Named Ranges ──────────────────────────────────────────────────────
    from openpyxl.workbook.defined_name import DefinedName

    # Revenue data
    dn = DefinedName("RevenueByRegion", attr_text="Dashboard!$A$17:$F$21")
    wb.defined_names.add(dn)

    dn2 = DefinedName("ActiveProjects", attr_text="Projects!$A$5:$H$13")
    wb.defined_names.add(dn2)

    dn3 = DefinedName("RiskMatrix", attr_text="'Risk Analysis'!$A$6:$G$13")
    wb.defined_names.add(dn3)

    # ─── Column widths ─────────────────────────────────────────────────────
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        ws.column_dimensions['A'].width = 30
        for col in range(2, 13):
            ws.column_dimensions[get_column_letter(col)].width = 15

    # Save
    output_path = os.path.join("excel_files", "complex_sample.xlsx")
    os.makedirs("excel_files", exist_ok=True)
    wb.save(output_path)
    print(f"Created: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Features: merged cells, multiple tables per sheet, charts, named ranges, instructions")
    return output_path


if __name__ == "__main__":
    create_complex_excel()
