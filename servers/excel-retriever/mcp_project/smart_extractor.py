"""
Intelligent Excel Extractor

Handles real-world Excel files with:
- Multiple tables per sheet (auto-detected)
- Merged cells
- Instruction text / notes mixed with data
- Figures and charts (metadata extraction)
- Named ranges
- Header detection and data type inference
- Empty row/column gaps between tables
"""

import json
import os
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field, asdict
from enum import Enum

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart


class ContentType(str, Enum):
    TABLE = "table"
    TEXT = "text"
    FIGURE = "figure"
    NAMED_RANGE = "named_range"
    MERGED_REGION = "merged_region"
    METADATA = "metadata"


@dataclass
class CellInfo:
    row: int
    col: int
    value: Any
    data_type: str
    is_merged: bool = False
    is_formula: bool = False
    number_format: str = ""
    font_bold: bool = False
    fill_color: Optional[str] = None


@dataclass
class DetectedTable:
    sheet_name: str
    table_index: int
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    headers: List[str]
    data: List[List[Any]]
    row_count: int
    col_count: int
    has_header: bool = True
    name: Optional[str] = None


@dataclass
class DetectedText:
    sheet_name: str
    row: int
    col: int
    text: str
    is_bold: bool = False
    classification: str = "note"  # "title", "instruction", "note", "label"


@dataclass
class DetectedFigure:
    sheet_name: str
    chart_type: str
    title: Optional[str]
    anchor: Optional[str]
    series_count: int = 0


@dataclass
class SheetAnalysis:
    sheet_name: str
    tables: List[DetectedTable] = field(default_factory=list)
    text_blocks: List[DetectedText] = field(default_factory=list)
    figures: List[DetectedFigure] = field(default_factory=list)
    merged_cells: List[Dict[str, Any]] = field(default_factory=list)
    named_ranges: List[Dict[str, Any]] = field(default_factory=list)
    total_rows: int = 0
    total_cols: int = 0


@dataclass
class TableRelationship:
    source_table: str
    target_table: str
    common_columns: List[str]
    relationship_type: str  # "exact_match", "overlap"
    description: str


@dataclass
class WorkbookAnalysis:
    file_path: str
    file_name: str
    sheet_count: int
    sheet_names: List[str]
    sheets: List[SheetAnalysis]
    named_ranges: List[Dict[str, Any]] = field(default_factory=list)
    relationships: List[TableRelationship] = field(default_factory=list)
    llm_schema_description: str = ""
    total_tables: int = 0
    total_figures: int = 0
    total_text_blocks: int = 0


def _extract_chart_title(title_obj) -> Optional[str]:
    """Extract plain text from an openpyxl chart Title object."""
    try:
        # Title is a string directly
        if isinstance(title_obj, str):
            return title_obj

        # Try to walk: title.tx.rich.paragraphs[].r[].t
        tx = getattr(title_obj, 'tx', None)
        if tx is not None:
            rich = getattr(tx, 'rich', None)
            if rich is not None:
                # rich is a RichText with .paragraphs (alias .p)
                paragraphs = getattr(rich, 'paragraphs', None) or getattr(rich, 'p', None) or []
                parts = []
                for para in paragraphs:
                    runs = getattr(para, 'r', []) or []
                    for run in runs:
                        t = getattr(run, 't', None)
                        if t:
                            parts.append(str(t))
                if parts:
                    return " ".join(parts)

        # Fallback: try .text attribute
        text = getattr(title_obj, 'text', None)
        if text and isinstance(text, str):
            return text
    except Exception:
        pass
    return None


class SmartExcelExtractor:
    """Intelligently extracts structured data from complex Excel files."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        # Also load with formulas for formula detection
        try:
            self.wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        except Exception:
            self.wb_formulas = None

    def analyze(self) -> WorkbookAnalysis:
        """Full workbook analysis: detect tables, text, figures, etc."""
        sheets = []
        total_tables = 0
        total_figures = 0
        total_text = 0

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            analysis = self._analyze_sheet(ws)
            sheets.append(analysis)
            total_tables += len(analysis.tables)
            total_figures += len(analysis.figures)
            total_text += len(analysis.text_blocks)

        # Extract workbook-level named ranges
        named_ranges = self._extract_named_ranges()

        # ─── New Knowledge-Enhancement Features ───
        
        # 1. Detect relationships between tables
        relationships = self._detect_relationships(sheets)
        
        # 2. Generate LLM-ready schema description
        llm_schema = self._generate_llm_schema(sheets, relationships, named_ranges)

        return WorkbookAnalysis(
            file_path=self.file_path,
            file_name=self.file_name,
            sheet_count=len(self.wb.sheetnames),
            sheet_names=list(self.wb.sheetnames),
            sheets=sheets,
            named_ranges=named_ranges,
            relationships=relationships,
            llm_schema_description=llm_schema,
            total_tables=total_tables,
            total_figures=total_figures,
            total_text_blocks=total_text,
        )

    def _detect_relationships(self, sheets: List[SheetAnalysis]) -> List[TableRelationship]:
        """Detect cross-references and foreign key relationships between tables."""
        all_tables = []
        for sheet in sheets:
            for table in sheet.tables:
                all_tables.append(table)
        
        relationships = []
        if len(all_tables) < 2:
            return relationships

        for i, table_a in enumerate(all_tables):
            name_a = f"{table_a.sheet_name}.{table_a.name or f'table_{table_a.table_index}'}"
            headers_a = set(h.lower() for h in table_a.headers)
            
            for j, table_b in enumerate(all_tables):
                if i >= j: continue
                
                name_b = f"{table_b.sheet_name}.{table_b.name or f'table_{table_b.table_index}'}"
                headers_b = set(h.lower() for h in table_b.headers)
                
                # Check for common columns
                common = headers_a.intersection(headers_b)
                # Filter out generic column names if they were generated
                common = {c for c in common if not c.startswith('column_')}
                
                if common:
                    # Found a relationship
                    common_list = sorted(list(common))
                    rel_type = "exact_match" if headers_a == headers_b else "overlap"
                    
                    description = f"Tables '{name_a}' and '{name_b}' share columns: {', '.join(common_list)}."
                    if rel_type == "exact_match":
                        description += " These tables appear to have identical schemas."
                    
                    relationships.append(TableRelationship(
                        source_table=name_a,
                        target_table=name_b,
                        common_columns=common_list,
                        relationship_type=rel_type,
                        description=description
                    ))
        
        return relationships

    def _generate_llm_schema(self, sheets: List[SheetAnalysis], 
                             relationships: List[TableRelationship],
                             named_ranges: List[Dict]) -> str:
        """Generate a natural language description of the workbook for LLM context."""
        lines = [f"Workbook: {self.file_name}", "=" * (len(self.file_name) + 10)]
        
        lines.append(f"\nThis workbook contains {len(sheets)} sheets and a total of {sum(len(s.tables) for s in sheets)} tables.")
        
        for sheet in sheets:
            lines.append(f"\n### Sheet: {sheet.sheet_name}")
            if not sheet.tables:
                lines.append("- No data tables found.")
            else:
                for table in sheet.tables:
                    table_name = table.name or f"Table {table.table_index}"
                    lines.append(f"- **{table_name}**: {table.row_count} rows x {table.col_count} columns.")
                    lines.append(f"  Columns: {', '.join(table.headers)}")
            
            if sheet.figures:
                lines.append(f"- Figures: {len(sheet.figures)} charts/images detected.")
            
            # Mention instructions if found
            instructions = [t for t in sheet.text_blocks if t.classification == "instruction"]
            if instructions:
                lines.append(f"- Contains {len(instructions)} instruction/note blocks.")

        if relationships:
            lines.append("\n### Key Relationships (Joins)")
            for rel in relationships:
                lines.append(f"- {rel.description}")

        if named_ranges:
            lines.append("\n### Named Ranges")
            for nr in named_ranges:
                lines.append(f"- '{nr['name']}' refers to {nr['value']}")

        lines.append("\n### AI Analysis Strategy")
        lines.append("1. Use common columns identified above to join data across tables.")
        lines.append("2. Pay attention to instruction blocks for data entry rules and business logic.")
        lines.append("3. References to 'ID' or 'Code' columns are likely primary/foreign keys.")

        return "\n".join(lines)

    def _analyze_sheet(self, ws) -> SheetAnalysis:
        """Analyze a single sheet for tables, text, figures."""
        sheet_name = ws.title
        analysis = SheetAnalysis(
            sheet_name=sheet_name,
            total_rows=ws.max_row or 0,
            total_cols=ws.max_column or 0,
        )

        if not ws.max_row or not ws.max_column:
            return analysis

        # 1. Extract merged cells info
        analysis.merged_cells = self._extract_merged_cells(ws)

        # 2. Build cell grid
        grid = self._build_cell_grid(ws)

        # 3. Detect figures/charts
        analysis.figures = self._extract_figures(ws)

        # 4. Detect tables (the core algorithm)
        analysis.tables = self._detect_tables(ws, grid, sheet_name)

        # 5. Extract standalone text blocks
        table_cells = set()
        for table in analysis.tables:
            for r in range(table.start_row, table.end_row + 1):
                for c in range(table.start_col, table.end_col + 1):
                    table_cells.add((r, c))

        analysis.text_blocks = self._extract_text_blocks(ws, grid, table_cells, sheet_name)

        return analysis

    def _build_cell_grid(self, ws) -> Dict[Tuple[int, int], CellInfo]:
        """Build a grid of cell information."""
        grid = {}
        merged_ranges = ws.merged_cells.ranges

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None or cell.coordinate in str(merged_ranges):
                    is_merged = any(
                        cell.coordinate in str(mr) for mr in merged_ranges
                    )
                    is_formula = False
                    if self.wb_formulas:
                        try:
                            fc = self.wb_formulas[ws.title][cell.coordinate]
                            if fc.value and isinstance(fc.value, str) and fc.value.startswith('='):
                                is_formula = True
                        except Exception:
                            pass

                    grid[(cell.row, cell.column)] = CellInfo(
                        row=cell.row,
                        col=cell.column,
                        value=cell.value,
                        data_type=cell.data_type,
                        is_merged=is_merged,
                        is_formula=is_formula,
                        number_format=cell.number_format or "",
                        font_bold=cell.font.bold if cell.font else False,
                        fill_color=str(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000' else None,
                    )
        return grid

    def _detect_tables(self, ws, grid: Dict, sheet_name: str) -> List[DetectedTable]:
        """
        Core algorithm: detect multiple tables on a sheet.

        Strategy:
        1. Find contiguous rectangular regions of non-empty cells.
        2. A table must have at least 2 rows and 2 columns.
        3. Tables are separated by empty rows/columns.
        4. Detect headers by checking if first row is text while others are mixed.
        """
        if not grid:
            return []

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Build occupancy matrix
        occupied = set()
        for (r, c), cell in grid.items():
            if cell.value is not None:
                occupied.add((r, c))

        visited = set()
        tables = []
        table_index = 0

        # Scan top-to-bottom, left-to-right for table anchors
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                if (r, c) in occupied and (r, c) not in visited:
                    # Try to expand into a table region
                    region = self._flood_fill_table(occupied, r, c, max_row, max_col)

                    if not region:
                        continue

                    # Mark all cells as visited
                    for cell_pos in region:
                        visited.add(cell_pos)

                    # Get bounding box
                    min_r = min(p[0] for p in region)
                    max_r_region = max(p[0] for p in region)
                    min_c = min(p[1] for p in region)
                    max_c_region = max(p[1] for p in region)

                    row_count = max_r_region - min_r + 1
                    col_count = max_c_region - min_c + 1

                    # Only consider as table if at least 2x2
                    if row_count >= 2 and col_count >= 2:
                        table_data = self._extract_table_data(
                            ws, min_r, max_r_region, min_c, max_c_region
                        )

                        # Detect if first row is a header
                        has_header, headers = self._detect_headers(
                            ws, grid, min_r, max_r_region, min_c, max_c_region
                        )

                        data_rows = table_data[1:] if has_header else table_data

                        # Try to find a table name from the row above
                        table_name = self._find_table_name(grid, min_r, min_c, max_c_region)

                        tables.append(DetectedTable(
                            sheet_name=sheet_name,
                            table_index=table_index,
                            start_row=min_r,
                            end_row=max_r_region,
                            start_col=min_c,
                            end_col=max_c_region,
                            headers=headers,
                            data=data_rows,
                            row_count=len(data_rows),
                            col_count=col_count,
                            has_header=has_header,
                            name=table_name,
                        ))
                        table_index += 1

        return tables

    def _flood_fill_table(self, occupied: set, start_r: int, start_c: int,
                          max_row: int, max_col: int) -> set:
        """
        Find a contiguous rectangular region starting from (start_r, start_c).

        Uses a row-based expansion: find the column extent of the first row,
        then expand downward as long as rows have similar occupancy.
        """
        # Find column extent of starting row
        col_start = start_c
        col_end = start_c

        # Expand right
        c = start_c + 1
        gap_count = 0
        while c <= max_col and gap_count < 2:
            if (start_r, c) in occupied:
                col_end = c
                gap_count = 0
            else:
                gap_count += 1
            c += 1

        # Need at least 2 columns for a table
        if col_end == col_start:
            return {(start_r, start_c)}

        # Now expand downward
        region = set()
        row_end = start_r

        for r in range(start_r, max_row + 1):
            # Count how many cells in this row (within col range) are occupied
            row_cells = sum(1 for c in range(col_start, col_end + 1) if (r, c) in occupied)
            col_span = col_end - col_start + 1

            # Row is part of table if >30% of columns are filled
            if row_cells > 0 and row_cells >= col_span * 0.3:
                row_end = r
                for c in range(col_start, col_end + 1):
                    region.add((r, c))
            elif r > start_r:
                # Allow 1 empty row gap within a table
                next_row_cells = sum(
                    1 for c in range(col_start, col_end + 1)
                    if (r + 1, c) in occupied
                ) if r + 1 <= max_row else 0

                if next_row_cells >= col_span * 0.3:
                    row_end = r
                    for c in range(col_start, col_end + 1):
                        region.add((r, c))
                else:
                    break
            else:
                break

        return region

    def _extract_table_data(self, ws, min_r: int, max_r: int,
                            min_c: int, max_c: int) -> List[List[Any]]:
        """Extract raw table data from the sheet."""
        data = []
        for r in range(min_r, max_r + 1):
            row_data = []
            for c in range(min_c, max_c + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                # Clean up value
                if val is None:
                    row_data.append(None)
                elif isinstance(val, str):
                    row_data.append(val.strip())
                else:
                    row_data.append(val)
            data.append(row_data)
        return data

    def _detect_headers(self, ws, grid: Dict, min_r: int, max_r: int,
                        min_c: int, max_c: int) -> Tuple[bool, List[str]]:
        """Detect if first row is a header row."""
        first_row = []
        first_row_all_text = True
        first_row_bold_count = 0

        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=min_r, column=c)
            val = cell.value
            first_row.append(str(val) if val is not None else f"Column_{c - min_c + 1}")

            if val is not None and not isinstance(val, str):
                first_row_all_text = False
            if cell.font and cell.font.bold:
                first_row_bold_count += 1

        col_count = max_c - min_c + 1

        # Header heuristics:
        # 1. All text in first row while data rows have numbers
        # 2. First row is bold
        # 3. First row has different formatting
        has_numeric_below = False
        if max_r > min_r:
            for c in range(min_c, max_c + 1):
                cell_below = ws.cell(row=min_r + 1, column=c)
                if isinstance(cell_below.value, (int, float)):
                    has_numeric_below = True
                    break

        is_header = (
            (first_row_all_text and has_numeric_below) or
            (first_row_bold_count >= col_count * 0.5) or
            (first_row_all_text and col_count >= 2)
        )

        headers = first_row if is_header else [
            f"Column_{i + 1}" for i in range(col_count)
        ]

        return is_header, headers

    def _find_table_name(self, grid: Dict, table_start_row: int,
                         min_col: int, max_col: int) -> Optional[str]:
        """Look for a table title in the row(s) above the table."""
        for offset in range(1, 3):
            check_row = table_start_row - offset
            if check_row < 1:
                break

            # Check if there's a single text cell spanning or near the table
            for c in range(min_col, max_col + 1):
                cell_info = grid.get((check_row, c))
                if cell_info and cell_info.value and isinstance(cell_info.value, str):
                    text = str(cell_info.value).strip()
                    # Likely a title if bold or relatively short
                    if cell_info.font_bold or (len(text) < 100 and not text.startswith('=')):
                        return text
        return None

    def _extract_text_blocks(self, ws, grid: Dict, table_cells: set,
                             sheet_name: str) -> List[DetectedText]:
        """Extract standalone text blocks that aren't part of tables."""
        text_blocks = []

        for (r, c), cell in grid.items():
            if (r, c) in table_cells:
                continue
            if cell.value is None:
                continue
            if not isinstance(cell.value, str):
                continue

            text = str(cell.value).strip()
            if not text:
                continue

            # Classify the text
            classification = "note"
            if cell.font_bold:
                if len(text) < 80:
                    classification = "title"
                else:
                    classification = "instruction"
            elif len(text) > 200:
                classification = "instruction"
            elif re.match(r'^(note|warning|important|caution|info)', text, re.IGNORECASE):
                classification = "instruction"

            text_blocks.append(DetectedText(
                sheet_name=sheet_name,
                row=r,
                col=c,
                text=text,
                is_bold=cell.font_bold,
                classification=classification,
            ))

        return text_blocks

    def _extract_figures(self, ws) -> List[DetectedFigure]:
        """Extract chart/figure metadata from the sheet."""
        figures = []

        for chart in ws._charts:
            chart_type = "unknown"
            if isinstance(chart, BarChart):
                chart_type = "bar"
            elif isinstance(chart, LineChart):
                chart_type = "line"
            elif isinstance(chart, PieChart):
                chart_type = "pie"
            elif isinstance(chart, AreaChart):
                chart_type = "area"
            elif isinstance(chart, ScatterChart):
                chart_type = "scatter"

            title = None
            if chart.title:
                title = _extract_chart_title(chart.title)

            series_count = len(chart.series) if hasattr(chart, 'series') else 0

            figures.append(DetectedFigure(
                sheet_name=ws.title,
                chart_type=chart_type,
                title=title,
                anchor=str(chart.anchor) if hasattr(chart, 'anchor') else None,
                series_count=series_count,
            ))

        # Also check for images
        if hasattr(ws, '_images'):
            for img in ws._images:
                figures.append(DetectedFigure(
                    sheet_name=ws.title,
                    chart_type="image",
                    title=None,
                    anchor=str(img.anchor) if hasattr(img, 'anchor') else None,
                ))

        return figures

    def _extract_merged_cells(self, ws) -> List[Dict[str, Any]]:
        """Extract merged cell regions."""
        merged = []
        for merge_range in ws.merged_cells.ranges:
            top_left = ws.cell(row=merge_range.min_row, column=merge_range.min_col)
            merged.append({
                "range": str(merge_range),
                "min_row": merge_range.min_row,
                "max_row": merge_range.max_row,
                "min_col": merge_range.min_col,
                "max_col": merge_range.max_col,
                "value": str(top_left.value) if top_left.value else None,
            })
        return merged

    def _extract_named_ranges(self) -> List[Dict[str, Any]]:
        """Extract workbook-level named ranges."""
        ranges = []
        try:
            for name_key in self.wb.defined_names:
                dn = self.wb.defined_names[name_key]
                destinations = []
                try:
                    for sheet_title, cell_range in dn.destinations:
                        destinations.append({
                            "sheet": sheet_title,
                            "range": cell_range,
                        })
                except Exception:
                    pass

                ranges.append({
                    "name": dn.name,
                    "value": str(dn.value) if dn.value else None,
                    "destinations": destinations,
                })
        except Exception:
            pass
        return ranges

    def close(self):
        """Close the workbook."""
        self.wb.close()
        if self.wb_formulas:
            self.wb_formulas.close()


# ─── Export Functions ───────────────────────────────────────────────────────────

def analysis_to_dict(analysis: WorkbookAnalysis) -> Dict[str, Any]:
    """Convert WorkbookAnalysis to a JSON-serializable dict."""
    result = {
        "file_path": analysis.file_path,
        "file_name": analysis.file_name,
        "sheet_count": analysis.sheet_count,
        "sheet_names": analysis.sheet_names,
        "total_tables": analysis.total_tables,
        "total_figures": analysis.total_figures,
        "total_text_blocks": analysis.total_text_blocks,
        "named_ranges": analysis.named_ranges,
        "relationships": [asdict(r) for r in analysis.relationships],
        "llm_schema_description": analysis.llm_schema_description,
        "sheets": [],
    }

    for sheet in analysis.sheets:
        sheet_dict = {
            "sheet_name": sheet.sheet_name,
            "total_rows": sheet.total_rows,
            "total_cols": sheet.total_cols,
            "tables": [],
            "text_blocks": [],
            "figures": [],
            "merged_cells": sheet.merged_cells,
        }

        for table in sheet.tables:
            sheet_dict["tables"].append({
                "table_index": table.table_index,
                "name": table.name,
                "start_row": table.start_row,
                "end_row": table.end_row,
                "start_col": table.start_col,
                "end_col": table.end_col,
                "headers": table.headers,
                "row_count": table.row_count,
                "col_count": table.col_count,
                "has_header": table.has_header,
                "data": table.data,
            })

        for text in sheet.text_blocks:
            sheet_dict["text_blocks"].append({
                "row": text.row,
                "col": text.col,
                "text": text.text,
                "is_bold": text.is_bold,
                "classification": text.classification,
            })

        for fig in sheet.figures:
            sheet_dict["figures"].append({
                "chart_type": fig.chart_type,
                "title": fig.title,
                "anchor": fig.anchor,
                "series_count": fig.series_count,
            })

        result["sheets"].append(sheet_dict)

    return result


def table_to_csv_string(table: DetectedTable) -> str:
    """Convert a detected table to CSV string."""
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)

    if table.has_header:
        writer.writerow(table.headers)

    for row in table.data:
        writer.writerow([str(v) if v is not None else "" for v in row])

    return output.getvalue()


def table_to_markdown(table: DetectedTable) -> str:
    """Convert a detected table to Markdown format."""
    lines = []

    # Header
    headers = table.headers
    lines.append("| " + " | ".join(str(h) for h in headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")

    # Data rows
    for row in table.data:
        cells = [str(v) if v is not None else "" for v in row]
        # Pad if needed
        while len(cells) < len(headers):
            cells.append("")
        lines.append("| " + " | ".join(cells[:len(headers)]) + " |")

    return "\n".join(lines)


def table_to_records(table: DetectedTable) -> List[Dict[str, Any]]:
    """Convert a detected table to list of dictionaries (records)."""
    records = []
    headers = table.headers

    for row in table.data:
        record = {}
        for i, header in enumerate(headers):
            record[header] = row[i] if i < len(row) else None
        records.append(record)

    return records
